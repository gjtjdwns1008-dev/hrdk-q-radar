import os
import json
import requests
import gspread
from datetime import datetime, timezone, timedelta
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 💡 1단계 config 파일에서 설정값들을 가져옵니다.
from config import COLUMNS, SUMMARY_COLUMNS, MAIN_SHEET_NAME, SYSTEM_NAME, WEBHOOK_URL, GCP_SERVICE_ACCOUNT_JSON, GOOGLE_SHEET_URL, TARGET_DATE

# 🌟 Track 코드 → 한글 병기 (구글 시트 표기용. SQLite엔 순수 코드 유지)
from hrdk_law_core.certs import label_track1_type, label_track1_risk, label_track2_code
from hrdk_law_core.certs import _normalize_cert
import re

# 시트에 병기로 표기할 Track 칸 — ★시트는 사람이 읽는 원장이므로 라벨 필수★
#   (brain은 맨 코드를 내고, 여기서 'B (영업요건형)'로 입혀 기록.
#    core 라벨 함수는 이미 라벨된 값을 그대로 통과시켜 이중 부착 위험 없음)
_TRACK_LABELERS = {
    "Track1_취급유형": label_track1_type,
    "Track1_위험도": label_track1_risk,
    "Track2_효용코드": label_track2_code,
}

def _row_for_sheet(info, columns):
    """COLUMNS 순서대로 행 구성 + Track 칸 한글 라벨 병기 (info 원본 불변)."""
    row = []
    for c in columns:
        val = info.get(c, "")
        labeler = _TRACK_LABELERS.get(c)
        row.append(labeler(val) if labeler else val)
    return row

# 🌟 [신설 헬퍼] 숫자를 엑셀 열 문자(1->A, 17->Q)로 변환해주는 함수
def get_column_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# ==========================================
# 1. 구글 시트 마스터 DB 적재 (통합 Upsert 엔진)
# ==========================================
# 🌟 [고도화] 관제용 상태(status)와 로그(log) 파라미터 기본값 추가
def _get_or_create_ws(spreadsheet, name, headers):
    """탭이 없으면 헤더와 함께 생성해 반환. (프레시 시트 대응 — Q-RADAR 신규)"""
    try:
        return spreadsheet.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=name, rows=200, cols=max(10, len(headers) + 2))
        ws.append_row(headers)
        print(f"  🆕 '{name}' 탭 생성 (헤더 {len(headers)}칸)")
        return ws


def _upsert_summary_row(spreadsheet, target_date_display, counts, symbol, log):
    """
    ★교훈의 코드화: 총괄현황표 기록 — 상태는 항상, 숫자는 성공 시에만.
      counts = [총, 연관높음, 단순, 우대] 또는 None(실패 — 건수 칸 불가침)
    상태 칸에는 "MM/DD HH:MM심볼"을 ' → '로 누적 (그날 시도 이력이 한 줄에 다 보임).
    """
    ws = _get_or_create_ws(spreadsheet, "총괄현황표", SUMMARY_COLUMNS)
    kst_now = datetime.now(timezone(timedelta(hours=9)))
    this_attempt = f"{kst_now.strftime('%m/%d %H:%M')}{symbol}"

    all_values = ws.get_all_values()
    header = all_values[0] if all_values else SUMMARY_COLUMNS
    idx = {h: i for i, h in enumerate(header)}
    st_i = idx.get("모니터링 상태", 5)
    lg_i = idx.get("실행 로그 및 비고", 6)

    row_num = None
    for r_i, row in enumerate(all_values[1:], start=2):
        if row and str(row[0]).strip() == target_date_display:
            row_num = r_i
            break

    def col(i):
        return get_column_letter(i + 1)

    if row_num is None:
        c = counts if counts is not None else ["", "", "", ""]
        new_row = [""] * len(header)
        new_row[0] = target_date_display
        for j, v in enumerate(c, start=1):
            if j < st_i:
                new_row[j] = v
        new_row[st_i] = this_attempt
        new_row[lg_i] = log
        ws.append_row(new_row)
    else:
        old_status = all_values[row_num - 1][st_i] if st_i < len(all_values[row_num - 1]) else ""
        new_status = f"{old_status} → {this_attempt}" if old_status.strip() else this_attempt
        updates = [{"range": f"{col(st_i)}{row_num}", "values": [[new_status]]},
                   {"range": f"{col(lg_i)}{row_num}", "values": [[log]]}]
        if counts is not None:
            updates.append({"range": f"B{row_num}:E{row_num}", "values": [counts]})
        ws.batch_update(updates)


def log_run_status(symbol, log, target_date=TARGET_DATE):
    """
    [공개 창구] 연결 실패 등 어떤 상황에서도 '상태만' 기록할 때 main이 부르는 함수.
    시트 인증부터 자체 처리하므로 파이프라인 어디서든 한 줄로 호출 가능.
    ※ 건수 칸은 절대 건드리지 않음 (숫자 불가침 원칙)
    """
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        spreadsheet = gspread.authorize(creds).open_by_key(GOOGLE_SHEET_URL)
        if target_date and len(str(target_date)) == 8:
            display_date = f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]}"
        else:
            display_date = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d")
        _upsert_summary_row(spreadsheet, display_date, None, symbol, log)
        print(f"  🧾 총괄현황표 상태 기록: {symbol}")
    except Exception as e:
        print(f"  ⚠️ 상태 기록 실패(치명 아님): {e}")


def fetch_main_ledger_values():
    """통합 대장 탭 전체 값 반환 (시트→SQLite 재구축용). 실패·미설정 시 빈 리스트."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return []
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        spreadsheet = gspread.authorize(creds).open_by_key(GOOGLE_SHEET_URL)
        return spreadsheet.worksheet(MAIN_SHEET_NAME).get_all_values()
    except gspread.WorksheetNotFound:
        return []
    except Exception as e:
        print(f"  ⚠️ 통합 대장 읽기 실패: {e}")
        return []


def upload_to_google_sheet(total_len, target_laws, target_date=TARGET_DATE, status="🟢 정상 작동", log="특이사항 없음"):
    """[Q-RADAR] 통합 대장(24칸) Upsert + 총괄현황표 기록(상태 항상·숫자 성공만)"""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        print("  ⚠️ 구글 시트 설정 정보가 없어 적재를 건너뜁니다.")
        return

    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)

        # ── 1) 총괄현황표: 통합 건수 4종 + 상태 누적 ──
        high = sum(1 for i in target_laws if i.get("연관도") == "연관높음")
        simple = sum(1 for i in target_laws if i.get("연관도") == "단순관련")
        preferred = sum(1 for i in target_laws if str(i.get("우대여부", "")).strip() == "O")
        try:
            if target_date and len(str(target_date)) == 8:
                display_date = f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]}"
            else:
                display_date = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d")
            symbol = "🟢"
            for s in ("🔴", "🟡", "🟢"):
                if s in status:
                    symbol = s
                    break
            counts = [total_len, high, simple, preferred] if symbol == "🟢" else None
            _upsert_summary_row(spreadsheet, display_date, counts, symbol, log)
        except Exception as e:
            print(f"  ⚠️ 총괄현황표 관제 데이터 기록 실패: {e}")

        # ── 2) 통합 대장 Upsert (자연키 = 법령명|시행일자 — 같은 법의 재개정은 별도 행) ──
        if target_laws:
            try:
                ws_main = _get_or_create_ws(spreadsheet, MAIN_SHEET_NAME, COLUMNS)
                existing_records = ws_main.get_all_records()

                max_id_num = 0
                natural_key_map = {}
                for idx, record in enumerate(existing_records):
                    mst_id = str(record.get("MST_ID", ""))
                    if mst_id.startswith("HRDK-L-"):
                        try:
                            num = int(mst_id.split("-")[-1])
                            if num > max_id_num:
                                max_id_num = num
                        except Exception:
                            pass
                    nat_key = f"{record.get('법령명','')}|{record.get('시행일자','')}"
                    natural_key_map[nat_key] = idx + 2

                new_rows_to_append = []
                end_col_letter = get_column_letter(len(COLUMNS))

                for info in target_laws:
                    nat_key = f"{info.get('법령명','')}|{info.get('시행일자','')}"
                    if nat_key in natural_key_map:
                        row_idx = natural_key_map[nat_key]
                        existing_id = existing_records[row_idx - 2].get("MST_ID", "")
                        info["MST_ID"] = existing_id
                        row_data = _row_for_sheet(info, COLUMNS)
                        ws_main.update(range_name=f'A{row_idx}:{end_col_letter}{row_idx}', values=[row_data])
                        print(f"  🔄 [Update] {existing_id}")
                    else:
                        max_id_num += 1
                        new_id = f"HRDK-L-{max_id_num:04d}"
                        info["MST_ID"] = new_id
                        row_data = _row_for_sheet(info, COLUMNS)
                        new_rows_to_append.append(row_data)
                        print(f"  ✨ [Insert] {new_id}")

                if new_rows_to_append:
                    ws_main.append_rows(new_rows_to_append)

            except Exception as e:
                print(f"  ⚠️ {MAIN_SHEET_NAME} 시트 적재 실패: {e}")

        print("  ✅ 구글 시트 통합 마스터 DB 적재 및 Upsert 완료!")

    except Exception as e:
        print(f"  ❌ 구글 시트 연동 중 에러: {e}")

# ==========================================
# 1-B. 보류목록 탭 내보내기 (코드가 채움, 담당자는 보기만)
# ==========================================
# 머리말(헤더): 담당자가 직관적으로 이해 + "검토상태"는 담당자가 직접 표기 가능
HELD_SHEET_NAME = "보류목록"
HELD_HEADERS = ["기록일시", "법령명", "시행일자", "소관부처", "보류사유", "법령링크", "검토상태"]


def export_held_laws_to_sheet(kb):
    """SQLite held_laws를 구글 시트 '보류목록' 탭으로 내보냅니다 (담당자 확인용)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)

        # 탭이 없으면 생성하고 헤더 작성
        try:
            ws = spreadsheet.worksheet(HELD_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=HELD_SHEET_NAME, rows=1000, cols=len(HELD_HEADERS))
            ws.append_row(HELD_HEADERS)

        # 미검토 보류건만 가져와 시트에 반영 (전체 덮어쓰기 대신 누적 append)
        held = kb.get_held_laws(only_unreviewed=True, limit=500)
        if not held:
            print("  ℹ️ 새로 보류된 법령 없음")
            return

        # 이미 시트에 있는 (법령명) 중복 방지
        existing = set()
        try:
            for r in ws.get_all_records():
                existing.add(str(r.get("법령명", "")))
        except Exception:
            pass

        new_rows = []
        for h in held:
            if h["law_name"] in existing:
                continue
            new_rows.append([
                h.get("created_at", ""), h["law_name"], h.get("enforce_date", ""),
                h.get("ministry", ""), h.get("hold_reason", ""), h.get("law_link", ""),
                "",  # 검토상태 — 담당자가 직접 기입
            ])
        if new_rows:
            ws.append_rows(new_rows)
            print(f"  📋 [보류목록] {len(new_rows)}건 시트 반영")
    except Exception as e:
        print(f"  ⚠️ 보류목록 시트 내보내기 실패: {e}")


# ==========================================
# 1-C. 자격명칭최신화 탭 (담당자가 편집 → 배치가 대장의 종목명을 실제 교체)
# ==========================================
# 이 탭은 '자격 명칭 변경' 작업 지시서입니다. 자격증 명칭이 바뀌면 담당자가 여기 적고,
# 배치가 변경시점이 지난 미적용 행을 읽어 대장(관련법령 탭)의 구 종목명을 신 종목명으로 교체합니다.
#   ※ 폐지/통합이라도 '자격 자체는 유효'하므로 삭제하지 않습니다. 오직 '명칭 교체'만 합니다.
UPDATE_SHEET_NAME = "자격명칭최신화"
# 헤더: 구명칭/신명칭/변경시점/적용여부/적용일시/비고
UPDATE_HEADERS = ["구명칭", "신명칭", "변경시점", "적용여부", "적용일시", "비고"]

# 예시 행(4개). 구명칭이 '[예시]'로 시작하는 행 + 그 위의 모든 행은 건너뜁니다.
# 즉 맨 아래 구분줄 다음부터가 실제 데이터입니다.
_UPDATE_EXAMPLES = [
    ["[예시] 전자계산기조직응용기사", "정보처리기사",   "2020-01-01", "",     "",                 "명칭 완전 변경: 구명칭→신명칭. 변경시점이 지나면 대장에서 교체"],
    ["[예시] 정보기기운용기능사",     "정보처리기능사", "2023-01-01", "",     "",                 "다른 자격과 합쳐지며 명칭이 바뀐 경우도 '구명칭→신명칭'으로 적으면 됨"],
    ["[예시] 미래에바뀔종목",         "새이름종목",     "2099-01-01", "",     "",                 "변경시점이 미래면 그날 전엔 적용 안 함(대기)"],
    ["[예시] 이미적용된예시",         "적용된신명칭",   "2020-01-01", "완료", "2026-01-01 00:00", "적용여부=완료 인 행은 다시 처리하지 않음"],
    ["", "", "", "", "", "═══ 실제 입력은 이 줄 아래부터 작성하세요 (위 [예시] 행들은 지우지 마세요) ═══"],
]


def ensure_update_sheet_exists():
    """자격명칭최신화 탭이 없으면 헤더 + 예시 행과 함께 생성합니다 (최초 1회용)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)
        try:
            spreadsheet.worksheet(UPDATE_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=UPDATE_SHEET_NAME, rows=1000, cols=len(UPDATE_HEADERS))
            ws.append_row(UPDATE_HEADERS)
            ws.append_rows(_UPDATE_EXAMPLES)
            print(f"  ✅ '{UPDATE_SHEET_NAME}' 탭 생성 (헤더 + 예시 {len(_UPDATE_EXAMPLES)-1}행 + 구분줄)")
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화 탭 확인 실패: {e}")


def read_all_aliases_for_resolve():
    """
    자격명칭최신화 탭에서 '분석 변환용' 별칭을 전부 읽어옵니다({구명칭: 신명칭}).
      · 대장 수정용(read_update_instructions)과 달리, 적용여부=완료 인 것도 포함.
        (과거 265개 이관분은 완료지만, 새 법령이 옛 명칭을 쓰면 변환해야 하므로 계속 사용)
      · 변경시점이 미래인 것만 제외(아직 발효 전 명칭변경은 변환에도 반영 안 함)
      · 예시행([예시]/구분줄 위쪽)은 건너뜀
    """
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return {}
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        ws = client.open_by_key(GOOGLE_SHEET_URL).worksheet(UPDATE_SHEET_NAME)
        values = ws.get_all_values()
        if len(values) <= 1:
            return {}
        header = values[0]
        idx = {h: i for i, h in enumerate(header)}
        def cell(row, name):
            i = idx.get(name)
            return (row[i].strip() if (i is not None and i < len(row)) else "")
        start = 1
        for r_i, row in enumerate(values[1:], start=1):
            gu = cell(row, "구명칭"); bigo = cell(row, "비고")
            if gu.startswith("[예시]") or "실제 입력은" in bigo or "═══" in bigo:
                start = r_i + 1
        today_digits = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")
        out = {}
        for r_i in range(start, len(values)):
            row = values[r_i]
            gu = cell(row, "구명칭"); sin = cell(row, "신명칭"); when = cell(row, "변경시점")
            if not gu or not sin or gu == sin:
                continue
            wd = "".join(ch for ch in when if ch.isdigit())[:8]
            if wd and len(wd) == 8 and wd > today_digits:  # 미래 발효분은 변환에도 아직 미반영
                continue
            out[gu] = sin
        return out
    except gspread.WorksheetNotFound:
        return {}
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화(변환용) 읽기 실패: {e}")
        return {}


def read_update_instructions():
    """
    자격명칭최신화 탭에서 '실제 적용할 명칭변경 지시'만 읽어옵니다.

    실제 데이터 시작점: 마지막 '[예시]' 행(또는 '═══ 실제 입력은...' 구분줄) '다음 줄'부터.
      → 예시/구분줄과 그 위쪽은 전부 건너뜀.

    각 지시 필터:
      - 구명칭·신명칭이 둘 다 있어야 함
      - 구명칭 == 신명칭이면 의미 없으므로 건너뜀
      - 적용여부가 이미 '완료'면 건너뜀 (1회성 실행 보장)
      - 변경시점이 오늘보다 미래면 건너뜀 (아직 발효 전)

    반환: [(row_num, 구명칭, 신명칭)] — row_num은 시트 행 번호(적용여부 기록용)
    """
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return []
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)
        ws = spreadsheet.worksheet(UPDATE_SHEET_NAME)
        values = ws.get_all_values()
        if len(values) <= 1:
            return []
        header = values[0]
        idx = {h: i for i, h in enumerate(header)}
        def cell(row, name):
            i = idx.get(name)
            return (row[i].strip() if (i is not None and i < len(row)) else "")

        # 실제 데이터 시작 행 찾기: 마지막 예시/구분줄의 다음 줄
        start = 1  # 0=헤더
        for r_i, row in enumerate(values[1:], start=1):
            gu = cell(row, "구명칭")
            bigo = cell(row, "비고")
            if gu.startswith("[예시]") or "실제 입력은" in bigo or "═══" in bigo:
                start = r_i + 1  # 이 줄 다음부터
        today_digits = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")

        out = []
        for r_i in range(start, len(values)):
            row = values[r_i]
            row_num = r_i + 1  # 시트는 1-based
            gu = cell(row, "구명칭")
            sin = cell(row, "신명칭")
            done = cell(row, "적용여부")
            when = cell(row, "변경시점")
            if not gu or not sin or gu == sin:
                continue
            if done:  # 이미 '완료' 등 표시가 있으면 재실행 안 함
                continue
            # 변경시점 발효 확인: 미래면 대기
            wd = "".join(ch for ch in when if ch.isdigit())[:8]
            if wd and len(wd) == 8 and wd > today_digits:
                continue
            out.append((row_num, gu, sin))
        return out
    except gspread.WorksheetNotFound:
        return []
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화 시트 읽기 실패: {e}")
        return []


def mark_update_applied(row_nums):
    """적용 완료한 지시 행의 '적용여부'=완료, '적용일시'=현재로 표시 (재실행 방지)."""
    if not row_nums or not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)
        ws = spreadsheet.worksheet(UPDATE_SHEET_NAME)
        header = ws.row_values(1)
        col_done = header.index("적용여부") + 1
        col_when = header.index("적용일시") + 1
        now = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")
        body = []
        for rn in row_nums:
            body.append({"range": gspread.utils.rowcol_to_a1(rn, col_done), "values": [["완료"]]})
            body.append({"range": gspread.utils.rowcol_to_a1(rn, col_when), "values": [[now]]})
        ws.batch_update(body, value_input_option="USER_ENTERED")
    except Exception as e:
        print(f"  ⚠️ 적용여부 기록 실패: {e}")


# ── 종목 문자열 안전 처리 (앞뒤 구분자 정리 + 중복 제거) ──
# 사전의 가운뎃점 종목명 보호 (build_site와 동일 원칙)
_DOT_CERTS = ["항공전기·전자정비기능사"]

def _split_cert_cell(raw):
    """종목 칸 문자열 → 종목 리스트. 괄호 안 쉼표 + 가운뎃점 종목명 보호."""
    s = str(raw or "")
    for dc in _DOT_CERTS:
        s = s.replace(dc, dc.replace("·", "㉿"))
    s = re.sub(r"\(([^)]*)\)", lambda m: "(" + m.group(1).replace(",", "§") + ")", s)
    return [c.strip().replace("§", ",").replace("㉿", "·") for c in re.split(r"[,/·\n]", s) if c.strip()]

def _join_cert_cell(items):
    """종목 리스트 → 칸 문자열 (표준 구분자 ', ')."""
    return ", ".join(items)

def apply_cert_updates_to_cell(cell_value, rename_map):
    """
    한 칸의 종목 문자열에 명칭변경(교체)을 적용.
      - rename_map: {정규화 구명칭: 신명칭}
    앞/중간/뒤 위치 무관하게 교체되고, 교체 후 같은 칸 내 중복은 하나만 남김.
    (자격 폐지/통합이라도 자격 자체는 유효하므로 '삭제'는 하지 않는다. 명칭 교체만.)
    반환: (새 칸 문자열, 변경여부)
    """
    items = _split_cert_cell(cell_value)
    if not items:
        return cell_value, False
    out, seen, changed = [], set(), False
    for it in items:
        key = _normalize_cert(it)
        if key in rename_map:          # 명칭변경 → 교체
            new = rename_map[key]
            if _normalize_cert(new) != key:
                changed = True
            it = new
        nk = _normalize_cert(it)
        if nk in seen:                 # 같은 칸 내 중복(교체 결과 등) → 하나만
            changed = True
            continue
        seen.add(nk)
        out.append(it)
    new_cell = _join_cert_cell(out)
    if new_cell != str(cell_value or "").strip():
        changed = True
    return new_cell, changed


def apply_name_updates_to_ledger(spreadsheet, sheet_name, cert_col_name,
                                 rename_map, preview=True):
    """
    대장 한 탭의 종목 칸에 명칭변경을 반영. preview=True면 미리보기만(실제 수정 안 함).
    반환: (변경된 행 수, 미리보기 목록[(행번호, 전, 후)])
    """
    try:
        ws = spreadsheet.worksheet(sheet_name)
    except Exception:
        print(f"    ⚠️ '{sheet_name}' 탭 없음 → 건너뜀")
        return 0, []
    values = ws.get_all_values()
    if len(values) <= 1:
        return 0, []
    header = values[0]
    if cert_col_name not in header:
        print(f"    ⚠️ '{sheet_name}'에 '{cert_col_name}' 칸 없음 → 건너뜀")
        return 0, []
    ci = header.index(cert_col_name)
    col_letter = gspread.utils.rowcol_to_a1(1, ci + 1).rstrip("1")

    changes, preview_list = [], []
    for r_i in range(1, len(values)):
        row = values[r_i]
        old = row[ci] if ci < len(row) else ""
        new, changed = apply_cert_updates_to_cell(old, rename_map)
        if changed and new != (old or "").strip():
            row_num = r_i + 1
            changes.append((row_num, new))
            preview_list.append((row_num, old, new))

    if not preview and changes:
        body = [{"range": f"{col_letter}{rn}", "values": [[nv]]} for rn, nv in changes]
        CHUNK = 500
        for i in range(0, len(body), CHUNK):
            ws.batch_update(body[i:i+CHUNK], value_input_option="USER_ENTERED")
    return len(changes), preview_list


# ==========================================
# 1-D. 우대사항 대장 탭 (법령+조문 단위 현황, 방식 B)
# ==========================================
LEDGER_SHEET_NAME = "우대사항_대장"
LEDGER_HEADERS = ["법령명", "조문", "우대분류", "해당 자격종목",
                  "Track1_취급유형", "Track1_위험도", "Track2_효용코드",
                  "중처법대상", "상태", "최근변경일", "비고"]


def _open_spreadsheet():
    """구글 시트 스프레드시트 객체 반환 (공통 인증)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return None
    creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    return client.open_by_key(GOOGLE_SHEET_URL)


def init_ledger_baseline(kb, resolve_fn=None):
    """
    [최초 1회] 우대사항 대장 기준선을 깝니다.
    이미 데이터가 있으면 건너뜁니다 (덮어쓰기 방지 = 방식 B).
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            return
        try:
            ws = ss.worksheet(LEDGER_SHEET_NAME)
            existing = ws.get_all_values()
            if len(existing) > 1:
                print(f"  ℹ️ 우대사항 대장에 이미 {len(existing)-1}행 존재 → 기준선 적재 생략")
                return
        except gspread.WorksheetNotFound:
            ws = ss.add_worksheet(title=LEDGER_SHEET_NAME, rows=5000, cols=len(LEDGER_HEADERS))
            ws.append_row(LEDGER_HEADERS)

        rows = kb.build_ledger_rows(resolve_fn=resolve_fn)
        # 배치로 한 번에 기록 (속도)
        data = [_row_for_sheet(r, LEDGER_HEADERS) for r in rows]
        if data:
            ws.append_rows(data)
            print(f"  📒 우대사항 대장 기준선 {len(data)}행 적재 완료")
    except Exception as e:
        print(f"  ⚠️ 우대사항 대장 기준선 적재 실패: {e}")


def fill_ledger_hazard_column(kb):
    """
    [일회성] 대장의 기존 행은 그대로 두고, '중처법대상' 칸만 채웁니다.
    (법령명+조문)으로 DB(build_ledger_rows)와 매칭하여 '대상'을 표시.
    기준선이 이미 깔린 뒤 중처법 칸을 추가한 경우 사용. 전체 덮어쓰기 없음.
    반환: 변경된 행 수.
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            print("  ⚠️ 스프레드시트 열기 실패")
            return 0
        ws = ss.worksheet(LEDGER_SHEET_NAME)
        records = ws.get_all_values()
        if len(records) <= 1:
            print("  ℹ️ 대장에 데이터가 없습니다.")
            return 0
        header = records[0]
        try:
            law_col = header.index("법령명")
            art_col = header.index("조문")
            hazard_col = header.index("중처법대상")
        except ValueError as e:
            print(f"  ⚠️ 대장 헤더에 필요한 칸이 없습니다: {e}")
            print(f"     (현재 헤더: {header})")
            return 0

        # DB에서 (법령명, 조문) → 중처법대상 매핑 생성
        rows = kb.build_ledger_rows()
        hazard_map = {}
        for r in rows:
            hazard_map[(r["법령명"], r["조문"])] = r.get("중처법대상", "")

        # 시트 각 행을 매칭해 중처법 칸 업데이트 대상 수집
        from gspread.utils import rowcol_to_a1
        updates = []
        filled = 0
        for i, row in enumerate(records[1:], start=2):
            law = row[law_col] if law_col < len(row) else ""
            art = row[art_col] if art_col < len(row) else ""
            cur = row[hazard_col] if hazard_col < len(row) else ""
            want = hazard_map.get((law, art), "")
            if want and want != cur:
                cell = rowcol_to_a1(i, hazard_col + 1)
                updates.append({"range": cell, "values": [[want]]})
                filled += 1

        if updates:
            ws.batch_update(updates)
            print(f"  ✅ 대장 중처법대상 {filled}개 행 채움 완료")
        else:
            print("  ℹ️ 채울 중처법대상 행이 없습니다 (이미 채워졌거나 매칭 없음).")
        return filled
    except Exception as e:
        print(f"  ⚠️ 대장 중처법대상 채우기 실패: {e}")
        return 0


def apply_cert_rename_to_ledger(old_name, new_name):
    """
    [명칭 변경 반영 - 방식 B] 대장에서 옛 종목명이 든 '해당 자격종목' 칸만
    찾아 현행명으로 교체합니다. 전체 덮어쓰기 없음 → 담당자 메모 보존.
    반환: 변경된 행 수.
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            return 0
        ws = ss.worksheet(LEDGER_SHEET_NAME)
        records = ws.get_all_values()
        if len(records) <= 1:
            return 0
        header = records[0]
        try:
            cert_col = header.index("해당 자격종목")
            chg_col = header.index("최근변경일")
        except ValueError:
            return 0

        from datetime import datetime, timezone, timedelta
        today = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d")
        changed = 0
        updates = []
        for i, row in enumerate(records[1:], start=2):
            if cert_col < len(row) and old_name in row[cert_col]:
                # 쉼표 구분 목록에서 정확히 그 종목만 치환
                certs = [c.strip() for c in row[cert_col].split(",")]
                new_certs = [new_name if c == old_name else c for c in certs]
                # 중복 제거
                seen, dedup = set(), []
                for c in new_certs:
                    if c not in seen:
                        seen.add(c); dedup.append(c)
                new_val = ", ".join(dedup)
                col_letter = get_column_letter(cert_col + 1)
                chg_letter = get_column_letter(chg_col + 1)
                updates.append({"range": f"{col_letter}{i}", "values": [[new_val]]})
                updates.append({"range": f"{chg_letter}{i}", "values": [[today]]})
                changed += 1
        if updates:
            ws.batch_update(updates)
            print(f"  🔤 대장 종목명 변경 반영: {old_name} → {new_name} ({changed}행)")
        return changed
    except gspread.WorksheetNotFound:
        return 0
    except Exception as e:
        print(f"  ⚠️ 대장 명칭 변경 반영 실패: {e}")
        return 0


# ==========================================
# 2. 엑셀 파일 생성 함수 (시트 1개로 단일화)
# ==========================================
UTIL_COLS = ["MST_ID", "시행일자", "소관부처", "법령명", "개정유형", "연관도",
             "관련 종목", "주요 제·개정내용", "활용도_구분", "활용도_상세",
             "근거조문", "조문별 다이렉트 링크"]
PREF_COLS = ["MST_ID", "시행일자", "법령명", "관련 종목", "우대분류",
             "Track1_취급유형", "Track1_위험도", "Track2_효용코드", "중처법대상",
             "조문 요약", "상세 분석 결과", "워크넷 실시간 구인건수", "조문별 다이렉트 링크"]


def create_excel_report(target_laws, target_date=TARGET_DATE, total_len=None):
    """[Q-RADAR] 일일 엑셀 3탭: 총괄현황표(대시보드) / 자격활용도분석 / 우대사항분석.
    각 탭이 자기 트랙의 컬럼만 보여준다 — 투트랙 시스템이 보고서 생김새에서 드러나게."""
    high = sum(1 for i in target_laws if i.get("연관도") == "연관높음")
    simple = sum(1 for i in target_laws if i.get("연관도") == "단순관련")
    prefs = [i for i in target_laws if str(i.get("우대여부", "")).strip() == "O"]

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "총괄현황표"
    ws0.append(["시행일자", "총 검토건수", "연관높음", "단순관련", "우대건수", "생성시각(KST)"])
    if target_date and len(str(target_date)) == 8:
        disp = f"{str(target_date)[:4]}-{str(target_date)[4:6]}-{str(target_date)[6:]}"
    else:
        disp = str(target_date)
    now_kst = datetime.now(timezone(timedelta(hours=9))).strftime("%m/%d %H:%M")
    ws0.append([disp, total_len if total_len is not None else "", high, simple, len(prefs), now_kst])
    for w, letter in zip([14, 12, 10, 10, 10, 16], "ABCDEF"):
        ws0.column_dimensions[letter].width = w

    ws1 = wb.create_sheet("자격활용도분석")
    ws1.append(UTIL_COLS)
    for info in target_laws:
        ws1.append(_row_for_sheet(info, UTIL_COLS))
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    ws2 = wb.create_sheet("우대사항분석")
    ws2.append(PREF_COLS)
    for info in prefs:
        ws2.append(_row_for_sheet(info, PREF_COLS))
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    excel_filename = f"HRDK_Q-RADAR_일일모니터링_{target_date}.xlsx"
    wb.save(excel_filename)
    return excel_filename

# ==========================================
# 3. 메이크닷컴 웹훅 전송 (기존과 동일)
# ==========================================
def send_webhook_with_file(fname, total, high, simple, preferred=0, target_date=TARGET_DATE):
    if not WEBHOOK_URL: return
    # 🌟 [근본 원인 해결!] 메일/웹훅으로 보낼 때도 사람이 읽기 편한 날짜로 변환해서 쏩니다!
    display_date = f"{target_date[:4]}년 {target_date[4:6]}월 {target_date[6:]}일"
    
    # 이제 Make.com은 "20260428"이 아니라 "2026년 04월 28일" 이라는 데이터를 받게 됩니다!
    # 🏷️ system/source: 두 시스템(RADAR/monitor)을 구분하는 식별값 (메일 제목 분기용)
    summary_data = {
        "system": SYSTEM_NAME,
        "source": "qradar",
        "subject": f"[Q-RADAR] {display_date} 법령 통합분석 (연관 {high}건·우대 {preferred}건)",
        "date": display_date, "total": f"{total}건", "high": f"{high}건",
        "simple": f"{simple}건", "preferred": f"{preferred}건"
    }
    try:
        if fname and os.path.exists(fname):
            with open(fname, 'rb') as f:
                requests.post(WEBHOOK_URL, data=summary_data, files={'file': (os.path.basename(fname), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
        else:
            requests.post(WEBHOOK_URL, data=summary_data)
        print("  ✅ 웹훅 전송 성공!")
    except Exception as e: print(f"  ❌ 웹훅 에러: {e}")
