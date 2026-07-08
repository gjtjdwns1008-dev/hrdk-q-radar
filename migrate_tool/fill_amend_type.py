# -*- coding: utf-8 -*-
"""
fill_amend_type.py — Q-RADAR 대장 정비 도구 (개정유형 백필 + 시행일자 교정)
=============================================================================
무엇을 하나요?
  ⓐ 개정유형이 빈 행(주로 RADAR 초기 구축기 유산 143행)을 법제처 목록 API의
     '제개정구분명'(제정/일부개정/전부개정/폐지 — AI 추측이 아닌 사실값)으로 채웁니다.
  ⓑ 시행일자에 섞인 '.0' 오염(예: 20140904.0)을 8자리로 교정합니다.
     (구글시트가 일부 셀을 숫자로 저장→내보내기에서 float화된 흔적. upsert 키를 어긋나게
      할 수 있어 반드시 청소)

건드리는 것: '국가기술자격 관련법령' 탭의 [시행일자]·[개정유형] 두 컬럼의 해당 셀만.
나머지 모든 셀·탭은 절대 불변.

실행: 그냥  python fill_amend_type.py  → 번호 메뉴 (1 미리보기 / 2 반영 / 0 종료)
준비물: migrate_tool/.env 에 QRADAR_SHEET + LAW_API_KEY, 폴더에 gcp-key.json
  ※ LAW_API_KEY가 .env에 없으면 메모장으로 한 줄 추가:  LAW_API_KEY=발급받은키
    (GitHub Secrets에 넣은 그 값과 동일)
"""
import os
import re
import sys
import time
import json
from pathlib import Path
import xml.etree.ElementTree as ET

HERE = Path(__file__).resolve().parent
PREVIEW_PATH = str(HERE / "대장정비_미리보기.xlsx")
MAIN_TAB = "국가기술자격 관련법령"


def _load_env_file():
    env = {}
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


_ENV = _load_env_file()


def norm_law(s):
    return re.sub(r"\s+", " ", str(s or "").strip())


def norm_date(v):
    d = "".join(ch for ch in str(v or "") if ch.isdigit())
    return d[:8] if len(d) >= 8 else ""


def _preflight():
    ok = True
    if not (HERE / "gcp-key.json").exists():
        print("🔑 gcp-key.json이 이 폴더에 없습니다 (이관 때 쓰던 그 파일).")
        ok = False
    if not _ENV.get("QRADAR_SHEET"):
        print("⚠️ .env에 QRADAR_SHEET가 없습니다 (이관 때 채운 .env 그대로면 있음).")
        ok = False
    if not _ENV.get("LAW_API_KEY"):
        print("⚠️ .env에 LAW_API_KEY가 없습니다.")
        print("   👉 메모장으로 .env를 열어 아래 한 줄을 추가하세요 (Secrets에 넣은 그 값):")
        print("      LAW_API_KEY=발급받은키")
        ok = False
    if not ok:
        print("\n준비물을 채운 뒤 다시 실행:  python fill_amend_type.py")
    return ok


def open_ws():
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    key = _ENV["QRADAR_SHEET"]
    m = re.search(r"/d/([A-Za-z0-9_-]+)", key)
    if m:
        key = m.group(1)
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        json.loads((HERE / "gcp-key.json").read_text(encoding="utf-8"), strict=False),
        ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"])
    ss = gspread.authorize(creds).open_by_key(key)
    return ss.worksheet(MAIN_TAB)


def _loose(s):
    """느슨 매칭용: 공백·「」·중점 제거 + 꼬리 괄호(약칭 등) 제거"""
    t = re.sub(r"[\s「」·ㆍ]", "", str(s or ""))
    t = re.sub(r"\([^)]*\)$", "", t)
    return t


def fetch_by_name(api_key, law_name):
    """2차 폴백: 법령명으로 직접 검색(law+histlaw) → [(명칭, 시행일자, 제개정구분명), …]"""
    import requests
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    out = []
    for target in ("law", "histlaw"):
        try:
            r = requests.get("https://www.law.go.kr/DRF/lawSearch.do",
                             params={"OC": api_key, "target": target, "type": "XML",
                                     "query": law_name, "display": 100},
                             headers=headers, timeout=30)
        except Exception:
            continue
        if r.status_code != 200 or not r.text.strip():
            continue
        try:
            root = ET.fromstring(r.text)
        except ET.ParseError:
            continue
        for law in root.findall(".//law"):
            nm = norm_law(law.findtext("법령명한글", ""))
            ef = norm_date(law.findtext("시행일자", ""))
            gbn = (law.findtext("제개정구분명", "") or law.findtext("제개정구분", "")).strip()
            if nm:
                out.append((nm, ef, gbn))
        time.sleep(0.2)
    return out


def fetch_amend_map(api_key, date8):
    """법제처 목록 API(law + histlaw)에서 해당 시행일자의 {법령명norm: 제개정구분명} 수집.
    core 스크레이퍼의 검증된 요청 패턴을 그대로 복제."""
    import requests
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    found = {}
    for target in ("law", "histlaw"):
        page = 1
        while page <= 10:                      # 안전 상한
            url = (f"https://www.law.go.kr/DRF/lawSearch.do"
                   f"?OC={api_key}&target={target}&type=XML"
                   f"&efYd={date8}~{date8}&display=100&page={page}")
            try:
                r = requests.get(url, headers=headers, timeout=30)
            except Exception:
                return found, False            # 통신 실패 → 이 날짜는 조회실패 처리
            if r.status_code != 200 or not r.text.strip():
                break
            try:
                root = ET.fromstring(r.text)
            except ET.ParseError:
                break
            nodes = root.findall(".//law")
            if not nodes:
                break
            for law in nodes:
                name = norm_law(law.findtext("법령명한글", ""))
                gbn = (law.findtext("제개정구분명", "") or law.findtext("제개정구분", "")).strip()
                if name and gbn and name not in found:
                    found[name] = gbn
            page += 1
            time.sleep(0.2)
    return found, True


def scan(ws):
    """대장 스캔 → (헤더, 시행일자 교정 목록, 개정유형 빈 행 목록)"""
    values = ws.get_all_values()
    if not values:
        print("⛔ 대장이 비어 있습니다."); sys.exit(1)
    header = [h.strip() for h in values[0]]
    ix = {h: i for i, h in enumerate(header)}
    for need in ("시행일자", "개정유형", "법령명", "MST_ID"):
        if need not in ix:
            print(f"⛔ 대장 헤더에 '{need}' 칸이 없습니다 — 탭/헤더 확인 필요."); sys.exit(1)

    date_fixes, amend_targets = [], []
    for rn, row in enumerate(values[1:], start=2):
        def cell(h):
            i = ix[h]
            return str(row[i]).strip() if i < len(row) else ""
        law = cell("법령명")
        if not law:
            continue
        raw_d, nd = cell("시행일자"), norm_date(cell("시행일자"))
        if nd and raw_d != nd:
            date_fixes.append((rn, raw_d, nd))
        if not cell("개정유형"):
            amend_targets.append({"row": rn, "mst": cell("MST_ID"), "date": nd or raw_d, "law": law})
    return header, ix, date_fixes, amend_targets


def run(apply: bool):
    print("=" * 62)
    print(f"🧰 Q-RADAR 대장 정비 — {'★실제 반영★' if apply else '미리보기만'}")
    print("=" * 62)
    ws = open_ws()
    header, ix, date_fixes, targets = scan(ws)
    print(f"  📥 대장 스캔 완료 — 시행일자 교정 대상 {len(date_fixes)}행 / 개정유형 빈 행 {len(targets)}행")

    # ── 법제처 조회 (날짜별 1세트: law+histlaw) ──
    dates = sorted({t["date"] for t in targets if len(t["date"]) == 8})
    print(f"  🌐 법제처 조회: {len(dates)}개 시행일자 (law+histlaw 이중 조회, 국내 IP)")
    api_key = _ENV["LAW_API_KEY"]
    date_map, failed_dates = {}, []
    for i, d in enumerate(dates, 1):
        found, ok = fetch_amend_map(api_key, d)
        date_map[d] = found
        if not ok:
            failed_dates.append(d)
        if i % 10 == 0 or i == len(dates):
            print(f"     …{i}/{len(dates)} 날짜 완료")

    fills, near_hits, not_found = [], [], []
    name_search_cache = {}
    pending_name_search = []

    for t in targets:
        dm = date_map.get(t["date"], {})
        lname = norm_law(t["law"])
        # 1단: 정확 일치
        if lname in dm:
            fills.append((t["row"], t["mst"], t["date"], t["law"], dm[lname], "정확"))
            continue
        # 2단: 느슨 일치 (공백·괄호·「」 차이)
        lloose = _loose(lname)
        loose_map = {_loose(k): (k, v) for k, v in dm.items()}
        if lloose and lloose in loose_map:
            apiname, gbn = loose_map[lloose]
            fills.append((t["row"], t["mst"], t["date"], t["law"], gbn, f"표기차이(API:{apiname})"))
            continue
        # 2.5단: 부분 포함 근사 — 자동 반영 제외, 수동확인 시트로
        cand = next(((k, v) for lk, (k, v) in loose_map.items()
                     if len(lloose) >= 6 and (lloose in lk or lk in lloose)), None)
        if cand:
            near_hits.append((t["row"], t["mst"], t["date"], t["law"], cand[1], f"근사(API:{cand[0]})"))
            continue
        # 3단: 명칭 직접 검색 폴백 (진단 겸용)
        pending_name_search.append(t)

    print(f"  🔎 명칭 직접 검색 폴백: {len(pending_name_search)}건")
    for i, t in enumerate(pending_name_search, 1):
        lname = norm_law(t["law"])
        if lname not in name_search_cache:
            name_search_cache[lname] = fetch_by_name(api_key, lname)
        cands = name_search_cache[lname]
        lloose = _loose(lname)
        same_name = [c for c in cands if _loose(c[0]) == lloose or
                     (len(lloose) >= 6 and (lloose in _loose(c[0]) or _loose(c[0]) in lloose))]
        exact_date = next((c for c in same_name if c[1] == t["date"] and c[2]), None)
        if exact_date:
            fills.append((t["row"], t["mst"], t["date"], t["law"], exact_date[2], "명칭검색-일자일치"))
        elif same_name:
            dates_seen = sorted({c[1] for c in same_name if c[1]})[:6]
            not_found.append((t["row"], t["mst"], t["date"], t["law"],
                              f"명칭은 존재하나 시행일자 불일치 — API상 시행이력: {', '.join(dates_seen)}"))
        else:
            reason = "조회 실패(통신)" if t["date"] in failed_dates else "명칭검색도 미발견 → 명칭 상이 또는 행정규칙(고시·훈령) 의심"
            not_found.append((t["row"], t["mst"], t["date"], t["law"], reason))
        if i % 20 == 0 or i == len(pending_name_search):
            print(f"     …명칭검색 {i}/{len(pending_name_search)}")

    from collections import Counter
    gbn_stat = Counter(f[4] for f in fills)
    how_stat = Counter(f[5] for f in fills)
    print(f"  ✅ 채울 수 있음: {len(fills)}행 {dict(gbn_stat)} | 매칭: {dict(how_stat)}")
    print(f"  🟡 근사(수동확인 필요, 자동 반영 안 함): {len(near_hits)}행")
    print(f"  ❓ 확인 불가: {len(not_found)}행 (사유별 목록 — 미리보기 참고)")

    # ── 미리보기 저장 ──
    from openpyxl import Workbook
    wb = Workbook()
    w1 = wb.active; w1.title = "개정유형_채움"
    w1.append(["행번호", "MST_ID", "시행일자", "법령명", "채울 값(법제처)", "매칭방식"])
    for f in fills:
        w1.append(list(f))
    wN = wb.create_sheet("근사후보(수동확인용)")
    wN.append(["행번호", "MST_ID", "시행일자", "법령명", "API 값(참고)", "비고 — 자동 반영 제외, 맞으면 직접 입력"])
    for nh in near_hits:
        wN.append(list(nh))
    w2 = wb.create_sheet("확인불가")
    w2.append(["행번호", "MST_ID", "시행일자", "법령명", "사유"])
    for nf in not_found:
        w2.append(list(nf))
    w3 = wb.create_sheet("시행일자_교정")
    w3.append(["행번호", "현재 값", "교정 값"])
    for df in date_fixes:
        w3.append(list(df))
    w4 = wb.create_sheet("통계")
    for k, v in [("개정유형 빈 행", len(targets)), ("채움 예정", len(fills)),
                 ("근사(수동확인)", len(near_hits)),
                 ("확인 불가", len(not_found)), ("시행일자 교정", len(date_fixes)),
                 ("조회 날짜 수", len(dates)), ("조회 실패 날짜", len(failed_dates)),
                 ("값 분포", str(dict(gbn_stat)))]:
        w4.append([k, v])
    try:
        wb.save(PREVIEW_PATH)
        print(f"  💾 미리보기 저장: {PREVIEW_PATH}")
    except PermissionError:
        alt = str(HERE / f"대장정비_미리보기_{time.strftime('%H%M%S')}.xlsx")
        wb.save(alt)
        print(f"  💾 미리보기 저장: {alt} (기존 파일이 엑셀에 열려 있음)")

    if not apply:
        print("\n✅ 미리보기만 생성했습니다. (시트는 아무것도 바뀌지 않았어요)")
        print("   👉 미리보기 확인 후, 다시 실행해서 2번(실제 반영)을 고르세요.")
        return

    # ── 실제 반영: 두 컬럼의 해당 셀만 ──
    from gspread.utils import rowcol_to_a1
    updates = []
    d_col = ix["시행일자"] + 1
    a_col = ix["개정유형"] + 1
    for rn, _raw, nd in date_fixes:
        updates.append({"range": rowcol_to_a1(rn, d_col), "values": [[nd]]})
    for rn, _m, _d, _l, gbn, _how in fills:
        updates.append({"range": rowcol_to_a1(rn, a_col), "values": [[gbn]]})
    if not updates:
        print("  반영할 것이 없습니다."); return
    for i in range(0, len(updates), 400):
        ws.batch_update(updates[i:i + 400], value_input_option="RAW")
        print(f"    …반영 {min(i + 400, len(updates))}/{len(updates)}셀")
    print(f"\n🎉 정비 완료! 시행일자 {len(date_fixes)}셀 교정 + 개정유형 {len(fills)}셀 채움.")
    print("   (확인 불가분은 빈칸 그대로 — 미리보기 '확인불가' 시트에 목록 있음)")


if __name__ == "__main__":
    args = [a.lower() for a in sys.argv[1:]]
    if args:                                   # 고급: preview/apply 인자도 지원
        run(apply=("apply" in args))
        sys.exit(0)
    print("=" * 62)
    print("🧰 Q-RADAR 대장 정비 도구 — 쉬운 모드")
    print("   (개정유형 백필 + 시행일자 .0 교정)")
    print("=" * 62)
    if not _preflight():
        sys.exit(0)
    print("\n무엇을 할까요?")
    print("  1) 미리보기 — 법제처 조회 결과로 '대장정비_미리보기.xlsx'만 생성 (안 바꿈)")
    print("  2) 실제 반영 — 시행일자·개정유형 셀만 채움/교정")
    print("  0) 종료")
    choice = input("\n번호를 입력하고 Enter: ").strip()
    if choice == "1":
        run(apply=False)
    elif choice == "2":
        confirm = input("⚠️ 시트의 두 컬럼 셀을 실제로 수정합니다. 진행하려면 '반영' 입력: ").strip()
        if confirm != "반영":
            print("취소했습니다. (아무 변화 없음)")
        else:
            run(apply=True)
    else:
        print("종료합니다. (아무 변화 없음)")
