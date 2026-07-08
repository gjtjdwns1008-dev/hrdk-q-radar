# -*- coding: utf-8 -*-
"""
backfill_usage.py — 활용도_구분 공란 백필 도구 (모드1: 분류 전용)
=============================================================================
무엇을 하나요?
  통합 대장에서 연관도가 [연관높음/단순관련]인데 '활용도_구분'만 빈 행을 찾아,
  이미 적혀 있는 '활용도_상세'와 '주요 제·개정내용'을 근거로 AI가 5단계
  (대폭 증가/소폭 증가/현상 유지/소폭 감소/대폭 감소) 중 하나를 골라
  **그 셀 하나만** 채웁니다. 나머지 23칸(수기 메모 포함)은 절대 건드리지 않습니다.

왜 생겼나요? (2026-07-05 정책 통일)
  과거엔 "연관높음일 때만 활용도"라는 게이트 탓에 단순관련 행의 구분이
  비어 있었습니다. brain.py 패치로 앞으로는 자동으로 채워지고,
  이 도구는 과거에 뚫린 구멍(약 73행)을 한 번에 메웁니다.

안전핀:
  · 1번(분석)은 시트를 읽기만 — 결과는 캐시(usage_cache.json)와
    미리보기 엑셀(활용도백필_미리보기.xlsx)로 저장. 눈으로 확인 후 반영.
  · 2번(반영)은 캐시 기반 — AI 재호출 없음(재과금 0). 기록 직전 행을
    재검증(여전히 공란인지, 연관도가 대상 범위인지)해 어긋나면 스킵+보고.
  · 세 칸이 모두 빈 행(약 47행)은 [4][5] 모드2로 처리 — 법제처 원문을
    로컬(국내 IP)에서 재수집해 3칸을 생성하고, "빈 셀만" 채웁니다.
  · 모든 네트워크 호출에 자동 재시도 3회(공단 PC 순간 컷 대비).

실행:  python backfill_usage.py   → 번호 메뉴 (1 분석+미리보기 / 2 시트 반영 / 3 대상 스캔만 / 0 종료)
준비물(.env): QRADAR_SHEET, GEMINI_API_KEY  + 폴더에 gcp-key.json
  + 로컬에 hrdk-law-core 설치 (로컬 백필 때 그 환경 그대로)
"""
import os
import re
import sys
import json
import time
from pathlib import Path

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
PREVIEW_PATH = str(HERE / "활용도백필_미리보기.xlsx")
CACHE_PATH = HERE / "usage_cache.json"
MAIN_TAB = "국가기술자격 관련법령"
TARGET_REL = ("연관높음", "단순관련")
FIVE = ("대폭 증가", "소폭 증가", "현상 유지", "소폭 감소", "대폭 감소")

COL_GUBUN = "활용도_구분"
COL_SANGSE = "활용도_상세"
COL_JUYO = "주요 제·개정내용"


# ── .env 로더 (reanalyze_ghosts 관례 계승) ─────────────────────────────
def _load_env_file():
    env = {}
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


_ENV = _load_env_file()
for _k in ("GEMINI_API_KEY", "LLM_PROVIDER", "LLM_MODEL"):
    if _ENV.get(_k) and not os.environ.get(_k):
        os.environ[_k] = _ENV[_k]

if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))


# ── 재시도 갑옷 (backfill_summary v2.1 계승) ────────────────────────────
def _retry(what, fn, tries=3, wait=4):
    for i in range(1, tries + 1):
        try:
            return fn()
        except Exception as e:
            if i == tries:
                raise
            print(f"  ⚠️ {what} 실패({i}/{tries}) — {str(e)[:60]} → {wait}초 후 재시도")
            time.sleep(wait)


# ── 시트 접속 ────────────────────────────────────────────────────────────
def open_ws():
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    key_path = HERE / "gcp-key.json"
    if not key_path.exists():
        print("❌ gcp-key.json이 이 폴더에 없습니다. (다른 migrate 도구와 같은 키 재사용)")
        sys.exit(1)
    sheet_ref = _ENV.get("QRADAR_SHEET", "").strip()
    if not sheet_ref:
        print("❌ .env에 QRADAR_SHEET가 없습니다.")
        sys.exit(1)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

    def _open():
        creds = ServiceAccountCredentials.from_json_keyfile_name(str(key_path), scope)
        gc = gspread.authorize(creds)
        ss = gc.open_by_url(sheet_ref) if sheet_ref.startswith("http") else gc.open_by_key(sheet_ref)
        return ss.worksheet(MAIN_TAB)

    return _retry("시트 접속", _open)


def _col_letter(n):  # 1→A, 27→AA
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# ── 대상 스캔 (순수 함수 — 오프라인 검증 가능) ──────────────────────────
def scan(records):
    """records: get_all_records() 결과.
    반환: (모드1 대상 목록, 모드2 후보 목록)
      모드1 = 연관도∈대상 & 구분 공란 & (상세 or 주요내용 있음)
      모드2 = 연관도∈대상 & 세 칸 모두 공란
    각 항목: {"row": 시트행번호, "mst", "date", "rel", "law", "sangse", "juyo"}"""
    t1, t2 = [], []
    for i, r in enumerate(records):
        rel = str(r.get("연관도", "") or "").strip()
        if rel not in TARGET_REL:
            continue
        g = str(r.get(COL_GUBUN, "") or "").strip()
        if g:
            continue
        item = {
            "row": i + 2,
            "mst": str(r.get("MST_ID", "") or "").strip(),
            "date": str(r.get("시행일자", "") or "").strip(),
            "rel": rel,
            "law": str(r.get("법령명", "") or "").strip(),
            "sangse": str(r.get(COL_SANGSE, "") or "").strip(),
            "juyo": str(r.get(COL_JUYO, "") or "").strip(),
            "certs": str(r.get("관련 종목", "") or "").strip(),
        }
        if item["sangse"] or item["juyo"]:
            t1.append(item)
        else:
            t2.append(item)
    return t1, t2


# ── AI 분류 ──────────────────────────────────────────────────────────────
def _normalize(raw):
    from brain import _normalize_usage
    return _normalize_usage(raw)


def ask_llm(prompt):
    from brain import _client
    return str(_client().generate_with_retry(prompt))


def _build_prompt(t, strict=False):
    lines = [
        "당신은 국가기술자격 정책 분석가입니다.",
        "아래 법령이 자격증 노동시장 활용도에 주는 변화를 다음 5개 중 정확히 하나로 분류하세요:",
        "대폭 증가 / 소폭 증가 / 현상 유지 / 소폭 감소 / 대폭 감소",
        "",
        f"[법령명] {t['law']}",
        f"[연관도] {t['rel']}",
    ]
    if t["juyo"]:
        lines.append(f"[주요 제·개정내용] {t['juyo'][:500]}")
    if t["sangse"]:
        lines.append(f"[활용도 분석] {t['sangse'][:500]}")
    lines += [
        "",
        "규칙: 근거 텍스트의 논조를 따르세요. 증가·감소 신호가 뚜렷하지 않으면 '현상 유지'.",
        "단순관련 법령은 간접 영향 관점에서 판단하세요.",
        "출력: 분류값 한 줄만. 설명·기호·다른 말 금지." + (" 반드시 5개 중 하나만 그대로 출력." if strict else ""),
    ]
    return "\n".join(lines)


def classify(t, ask=ask_llm):
    """5택1 분류. 1차 실패 시 강경 문구로 1회 재시도. 실패면 ""."""
    for strict in (False, True):
        try:
            raw = ask(_build_prompt(t, strict))
        except Exception as e:
            print(f"    ⚠️ LLM 호출 실패({t['mst']}): {str(e)[:60]}")
            time.sleep(2)
            continue
        cls = _normalize(raw)
        if cls in FIVE:
            return cls
    return ""


def run_classify(targets, ask=ask_llm, cache=None, sleep=0.4):
    """대상 목록 분류. 캐시에 이미 유효 분류가 있으면 재호출 없이 재사용(이어달리기)."""
    cache = cache if cache is not None else {}
    reused = 0
    for k, t in enumerate(targets, 1):
        hit = cache.get(t["mst"], {})
        if hit.get("cls") in FIVE:
            reused += 1
            print(f"  [{k}/{len(targets)}] {t['mst']} {t['law'][:24]} → {hit['cls']} (캐시)")
            continue
        cls = classify(t, ask=ask)
        cache[t["mst"]] = {"cls": cls, "law": t["law"], "date": t["date"], "rel": t["rel"], "row": t["row"]}
        mark = cls if cls else "❌ 분류 실패"
        print(f"  [{k}/{len(targets)}] {t['mst']} {t['law'][:24]} → {mark}")
        if k % 10 == 0:
            _save_cache(cache)
        time.sleep(sleep)
    _save_cache(cache)
    if reused:
        print(f"  ↻ 캐시 재사용 {reused}건 (API 재과금 없음)")
    return cache


def _save_cache(cache):
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=1), encoding="utf-8")


def _load_cache():
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


# ── 미리보기 엑셀 ────────────────────────────────────────────────────────
def write_preview(targets, cache, mode2):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "분류결과(반영 예정)"
    ws.append(["MST_ID", "시행일자", "연관도", "법령명", "→ 활용도_구분", "근거 발췌"])
    ok = fail = 0
    for t in targets:
        cls = cache.get(t["mst"], {}).get("cls", "")
        basis = (t["sangse"] or t["juyo"])[:60]
        ws.append([t["mst"], t["date"], t["rel"], t["law"], cls if cls else "❌ 실패(재실행 시 재시도)", basis])
        ok, fail = (ok + 1, fail) if cls else (ok, fail + 1)
    ws2 = wb.create_sheet("모드2 대기(3칸 공란)")
    ws2.append(["MST_ID", "시행일자", "연관도", "법령명", "비고"])
    for t in mode2:
        ws2.append([t["mst"], t["date"], t["rel"], t["law"], "원문 재수집 필요 — 이 도구 범위 밖"])
    wb.save(PREVIEW_PATH)
    return ok, fail


# ── 시트 반영 (2번 메뉴 핵심 — 셀 단위·재검증) ──────────────────────────
def apply_core(ws, cache):
    """캐시의 분류값을 시트에 기록. 기록 직전 재검증:
      · MST 일치 행이 존재하고
      · 그 행의 활용도_구분이 여전히 공란이며
      · 연관도가 대상 범위일 때만.
    반환: (기록 n, 스킵 목록)"""
    header = _retry("헤더 읽기", lambda: ws.row_values(1))
    try:
        g_idx = header.index(COL_GUBUN) + 1
    except ValueError:
        print(f"❌ 시트에 '{COL_GUBUN}' 열이 없습니다."); sys.exit(1)
    col_l = _col_letter(g_idx)

    records = _retry("대장 읽기", lambda: ws.get_all_records())
    by_mst = {}
    for i, r in enumerate(records):
        m = str(r.get("MST_ID", "") or "").strip()
        if m:
            by_mst[m] = (i + 2, r)

    updates, skips = [], []
    for mst, info in cache.items():
        cls = info.get("cls", "")
        if cls not in FIVE:
            skips.append((mst, "분류 실패분 — 1번 재실행으로 채운 뒤 반영"))
            continue
        if mst not in by_mst:
            skips.append((mst, "시트에 해당 MST 없음"))
            continue
        row, r = by_mst[mst]
        if str(r.get(COL_GUBUN, "") or "").strip():
            skips.append((mst, f"이미 값 있음('{str(r.get(COL_GUBUN)).strip()}') — 보존"))
            continue
        if str(r.get("연관도", "") or "").strip() not in TARGET_REL:
            skips.append((mst, "연관도가 대상 범위 밖으로 변경됨"))
            continue
        updates.append({"range": f"{col_l}{row}", "values": [[cls]]})

    done = 0
    for i in range(0, len(updates), 20):
        chunk = updates[i:i + 20]
        _retry(f"기록 {i + 1}~{i + len(chunk)}", lambda c=chunk: ws.batch_update(c))
        done += len(chunk)
        print(f"  ✍️ {done}/{len(updates)}칸 기록")
    return done, skips



# ── 모드2: 3칸 공란 행 — 법제처 재수집 + 3칸 생성 ───────────────────────
CACHE2_PATH = HERE / "usage_cache_mode2.json"
PREVIEW2_PATH = str(HERE / "활용도백필_모드2_미리보기.xlsx")


def _law_api_key():
    k = _ENV.get("LAW_API_KEY", "") or os.environ.get("LAW_API_KEY", "")
    if not k:
        print("❌ .env에 LAW_API_KEY가 없습니다 (모드2는 법제처 재수집이 필요)")
        sys.exit(1)
    return k


def fetch_laws_by_date(date):
    # core 수집기 재사용 — 일일 파이프라인과 동일 경로 (로컬 국내 IP라 차단 걱정 없음)
    from hrdk_law_core.scraper import get_base_laws
    return get_base_laws(api_key=_law_api_key(), target_date=date) or []


def _nsp(x):
    # 공백 + 사잇점 계열(ㆍ·・‧) 제거 — 직능연 옛 표기(사잇점 없음)와
    # 법제처 공식명(사잇점 있음)을 같은 이름으로 매칭하기 위함
    return re.sub(r"[\s\u318D\u00B7\u30FB\u2027]+", "", str(x or ""))


def _build_prompt2(t, src, strict=False):
    body = str(src.get("원본", ""))[:3500]
    lines = [
        "당신은 국가기술자격 정책 분석가입니다. 아래 법령 원문을 읽고 세 항목을 JSON으로 작성하세요.",
        "",
        "[법령명] " + t["law"],
        "[소관부처] " + str(src.get("소관부처", "")),
        "[시행일자] " + t["date"],
        ("[관련 자격증] " + t.get("certs", "")) if t.get("certs") else "",
        "",
        "[원문 발췌]",
        body if body else "(본문 없음 — 법령명과 자격증 맥락으로 보수적으로 판단)",
        "",
        '출력(JSON만, 코드펜스·설명 금지):',
        '{"주요_제개정내용": "- 팩트1 - 팩트2 (글머리 - 나열, 한 줄)",',
        ' "활용도_구분": "대폭 증가/소폭 증가/현상 유지/소폭 감소/대폭 감소 중 택1",',
        ' "활용도_상세": "자격 수요·활용도 관점 3문장 이내 (주요_제개정내용과 반복 금지)"}',
        "규칙: 증가·감소 근거가 뚜렷하지 않으면 '현상 유지'."
        + (" JSON 외 어떤 문자도 출력하지 마세요." if strict else ""),
    ]
    return "\n".join(x for x in lines if x != "")


def _parse_json(raw):
    m = re.search(r"\{.*\}", str(raw), re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None


def classify3(t, src, ask=ask_llm):
    # 3칸 생성. 1차 실패 시 강경 문구로 1회 재시도. 실패면 ("","","").
    for strict in (False, True):
        try:
            raw = ask(_build_prompt2(t, src, strict))
        except Exception as e:
            print(f"    ⚠️ LLM 호출 실패({t['mst']}): {str(e)[:60]}")
            time.sleep(2)
            continue
        d = _parse_json(raw)
        if not d:
            continue
        g = _normalize(d.get("활용도_구분", ""))
        ju = str(d.get("주요_제개정내용", "")).strip()
        sa = str(d.get("활용도_상세", "")).strip()
        if g in FIVE and ju and sa:
            return ju, g, sa
    return "", "", ""


def _save2(cache):
    CACHE2_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=1), encoding="utf-8")


def _load2():
    if CACHE2_PATH.exists():
        try:
            return json.loads(CACHE2_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def run_mode2(targets, fetch=None, ask=ask_llm, cache=None, sleep=0.4):
    # 시행일자별로 법제처를 1회씩 수집(캐시 완료 날짜는 건너뜀) → 법령명 매칭 → 3칸 생성
    fetch = fetch or fetch_laws_by_date
    cache = cache if cache is not None else {}

    def _done(v):
        return v.get("g") in FIVE and v.get("ju") and v.get("sa")

    pending = [t for t in targets if not _done(cache.get(t["mst"], {}))]
    reused = len(targets) - len(pending)
    if reused:
        print(f"  ↻ 캐시 재사용 {reused}건 (수집·API 재과금 없음)")
    dates = sorted({t["date"] for t in pending})
    print(f"  📅 수집할 시행일자 {len(dates)}개 / 생성 대상 {len(pending)}건")

    law_idx = {}
    for i, d in enumerate(dates, 1):
        try:
            laws = _retry(f"법제처 수집 {d}", lambda dd=d: fetch(dd))
        except Exception as e:
            print(f"  [{i}/{len(dates)}] {d}: ❌ 수집 실패 — {str(e)[:50]}")
            laws = []
        law_idx[d] = {_nsp(l.get("법령명", "")): l for l in (laws or [])}
        print(f"  [{i}/{len(dates)}] {d}: {len(laws or [])}건 수집")

    for k, t in enumerate(pending, 1):
        src = law_idx.get(t["date"], {}).get(_nsp(t["law"]))
        if not src:
            cache[t["mst"]] = {"ju": "", "g": "", "sa": "", "law": t["law"], "date": t["date"],
                               "err": "법제처 미발견(해당 시행일 목록에 없음)"}
            print(f"  [{k}/{len(pending)}] {t['mst']} {t['law'][:22]} → ❌ 법제처 미발견")
            _save2(cache)
            continue
        ju, g, sa = classify3(t, src, ask=ask)
        cache[t["mst"]] = {"ju": ju, "g": g, "sa": sa, "law": t["law"], "date": t["date"],
                           "err": "" if g else "생성 실패"}
        print(f"  [{k}/{len(pending)}] {t['mst']} {t['law'][:22]} → {g if g else '❌ 생성 실패'}")
        if k % 5 == 0:
            _save2(cache)
        time.sleep(sleep)
    _save2(cache)
    return cache


def write_preview2(targets, cache):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "모드2 생성결과(반영 예정)"
    ws.append(["MST_ID", "시행일자", "연관도", "법령명", "→ 주요 제·개정내용", "→ 활용도_구분", "→ 활용도_상세", "비고"])
    ok = fail = 0
    for t in targets:
        v = cache.get(t["mst"], {})
        good = v.get("g") in FIVE and v.get("ju") and v.get("sa")
        ws.append([t["mst"], t["date"], t["rel"], t["law"],
                   v.get("ju", "")[:120], v.get("g", ""), v.get("sa", "")[:120],
                   "" if good else (v.get("err") or "생성 실패 — [4] 재실행 시 재시도")])
        ok, fail = (ok + 1, fail) if good else (ok, fail + 1)
    wb.save(PREVIEW2_PATH)
    return ok, fail


def apply2_core(ws, cache):
    # 3칸 각각 "지금 비어 있는 셀만" 기록 — 값 있는 셀은 무조건 보존.
    header = _retry("헤더 읽기", lambda: ws.row_values(1))
    letters = {}
    for cname in (COL_JUYO, COL_GUBUN, COL_SANGSE):
        try:
            letters[cname] = _col_letter(header.index(cname) + 1)
        except ValueError:
            print(f"❌ 시트에 '{cname}' 열이 없습니다.")
            sys.exit(1)

    records = _retry("대장 읽기", lambda: ws.get_all_records())
    by_mst = {}
    for i, r in enumerate(records):
        m = str(r.get("MST_ID", "") or "").strip()
        if m:
            by_mst[m] = (i + 2, r)

    updates, skips, rows_touched = [], [], set()
    for mst, v in cache.items():
        if not (v.get("g") in FIVE and v.get("ju") and v.get("sa")):
            skips.append((mst, v.get("err") or "생성 실패분 — [4] 재실행으로 채운 뒤 반영"))
            continue
        if mst not in by_mst:
            skips.append((mst, "시트에 해당 MST 없음"))
            continue
        row, r = by_mst[mst]
        if str(r.get("연관도", "") or "").strip() not in TARGET_REL:
            skips.append((mst, "연관도가 대상 범위 밖으로 변경됨"))
            continue
        wrote = False
        for cname, val in ((COL_JUYO, v["ju"]), (COL_GUBUN, v["g"]), (COL_SANGSE, v["sa"])):
            if str(r.get(cname, "") or "").strip():
                continue  # 이미 값 있음 → 보존
            updates.append({"range": f"{letters[cname]}{row}", "values": [[val]]})
            wrote = True
        if wrote:
            rows_touched.add(mst)
        else:
            skips.append((mst, "3칸 모두 이미 값 있음 — 보존"))

    done = 0
    for i in range(0, len(updates), 20):
        chunk = updates[i:i + 20]
        _retry(f"기록 {i + 1}~{i + len(chunk)}", lambda c=chunk: ws.batch_update(c))
        done += len(chunk)
        print(f"  ✍️ {done}/{len(updates)}칸 기록")
    return done, len(rows_touched), skips


# ── 메뉴 ─────────────────────────────────────────────────────────────────
def _print_scan(t1, t2):
    print(f"\n📋 스캔 결과: 모드1 대상(구분만 공란) {len(t1)}건 / 모드2 대기(3칸 공란) {len(t2)}건")
    for t in t1[:5]:
        print(f"   · {t['mst']} [{t['rel']}] {t['law'][:30]}")
    if len(t1) > 5:
        print(f"   … 외 {len(t1) - 5}건")


def main():
    print("=" * 62)
    print("  활용도_구분 백필 도구 (모드1: 분류 전용)  —  HRDK Q-RADAR")
    print("=" * 62)
    while True:
        print("\n[1] 모드1 분석+미리보기 (구분만 공란 — 근거 기반 분류)")
        print("[2] 모드1 시트 반영")
        print("[3] 대상 스캔만 (AI·기록 없음)")
        print("[4] 모드2 분석+미리보기 (3칸 공란 — 법제처 재수집+생성)")
        print("[5] 모드2 시트 반영")
        print("[0] 종료")
        sel = input("번호 선택: ").strip()

        if sel == "0":
            print("종료합니다."); return

        if sel == "3":
            ws = open_ws()
            t1, t2 = scan(_retry("대장 읽기", lambda: ws.get_all_records()))
            _print_scan(t1, t2)

        elif sel == "1":
            ws = open_ws()
            t1, t2 = scan(_retry("대장 읽기", lambda: ws.get_all_records()))
            _print_scan(t1, t2)
            if not t1:
                print("✅ 채울 대상이 없습니다!"); continue
            if input(f"\n{len(t1)}건 AI 분류를 시작할까요? (y/n): ").strip().lower() != "y":
                continue
            cache = run_classify(t1, cache=_load_cache())
            ok, fail = write_preview(t1, cache, t2)
            print(f"\n✅ 분류 {ok}건 / 실패 {fail}건")
            print(f"📄 미리보기: {PREVIEW_PATH}")
            print("   → 엑셀 확인 후 [2]번으로 반영하세요." + (" 실패분은 [1]번 재실행 시 그 행만 재시도됩니다." if fail else ""))

        elif sel == "2":
            cache = _load_cache()
            valid = sum(1 for v in cache.values() if v.get("cls") in FIVE)
            if not valid:
                print("⚠️ 캐시가 비어 있습니다. 먼저 [1]번을 실행하세요."); continue
            print(f"캐시에 유효 분류 {valid}건. 미리보기 엑셀을 확인하셨나요?")
            if input("시트에 기록할까요? (y/n): ").strip().lower() != "y":
                continue
            ws = open_ws()
            done, skips = apply_core(ws, cache)
            print(f"\n✅ 기록 완료: {done}칸 (활용도_구분 셀만)")
            if skips:
                print(f"⏭️ 스킵 {len(skips)}건:")
                for m, why in skips[:10]:
                    print(f"   · {m}: {why}")
                if len(skips) > 10:
                    print(f"   … 외 {len(skips) - 10}건")
        elif sel == "4":
            ws = open_ws()
            t1, t2 = scan(_retry("대장 읽기", lambda: ws.get_all_records()))
            print(f"\n📋 모드2 대상(3칸 공란): {len(t2)}건")
            if not t2:
                print("✅ 채울 대상이 없습니다!"); continue
            if input(f"법제처 재수집 + AI 생성을 시작할까요? (y/n): ").strip().lower() != "y":
                continue
            cache = run_mode2(t2, cache=_load2())
            ok, fail = write_preview2(t2, cache)
            print(f"\n✅ 생성 {ok}건 / 실패 {fail}건")
            print(f"📄 미리보기: {PREVIEW2_PATH}")
            print("   → 엑셀 확인 후 [5]번으로 반영하세요.")

        elif sel == "5":
            cache = _load2()
            valid = sum(1 for v in cache.values() if v.get("g") in FIVE and v.get("ju") and v.get("sa"))
            if not valid:
                print("⚠️ 모드2 캐시가 비어 있습니다. 먼저 [4]번을 실행하세요."); continue
            print(f"모드2 캐시에 유효 생성 {valid}건. 미리보기 엑셀을 확인하셨나요?")
            if input("시트에 기록할까요? (y/n): ").strip().lower() != "y":
                continue
            ws = open_ws()
            done, rows_n, skips = apply2_core(ws, cache)
            print(f"\n✅ 기록 완료: {rows_n}행 / {done}칸 (빈 셀만)")
            if skips:
                print(f"⏭️ 스킵 {len(skips)}건:")
                for m, why in skips[:10]:
                    print(f"   · {m}: {why}")
                if len(skips) > 10:
                    print(f"   … 외 {len(skips) - 10}건")

        else:
            print("1~5 또는 0 중에서 선택하세요.")


if __name__ == "__main__":
    main()
