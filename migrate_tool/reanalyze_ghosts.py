# -*- coding: utf-8 -*-
"""
reanalyze_ghosts.py — Q-RADAR '기준선 유령' 재분석 도구 (v3)
=============================================================================
무엇을 하나요?
  개정유형이 빈 채 남은 행(직능연 2022 기준선 유령 + 개명 법령)을,
  ⓐ 개명 6종 매핑으로 현행 법령명을 확정하고
  ⓑ 법제처에서 현행법 원문을 받아 통합 brain으로 '오늘 기준' 재분석하여
  ⓒ 그 행 자리에 in-place 갱신합니다 (MST_ID 유지, 이력은 검토사유에 기록).

안전핀:
  · 미리보기 먼저 — 어떤 행이 어떤 현행법으로 매칭되는지, 키 충돌은 없는지 확인 후 반영
  · 현행 시행일자 키가 대장에 이미 있으면(일일 수집분 존재) 그 행은 스킵+보고
  · 직능연 2022 원본은 '우대사항_대장' 탭이 계속 보존 — 여긴 일일 대장의 현행화

실행:  python reanalyze_ghosts.py   → 번호 메뉴 (1 미리보기 / 2 재분석 반영 / 0 종료)
준비물(.env): QRADAR_SHEET, LAW_API_KEY, GEMINI_API_KEY (+선택 WORKNET_API_KEY)
  + 폴더에 gcp-key.json, 로컬에 hrdk-law-core 설치(로컬 백필 때 그 환경 그대로)
"""
import os
import re
import sys
import time
import json
from pathlib import Path
import xml.etree.ElementTree as ET

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
PREVIEW_PATH = str(HERE / "재분석_미리보기.xlsx")
MAIN_TAB = "국가기술자격 관련법령"


def _load_env_file():
    env = {}
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


_ENV = _load_env_file()
# 레포 모듈(brain·config)과 core가 환경변수를 읽으므로 가장 먼저 주입
for _k in ("LAW_API_KEY", "GEMINI_API_KEY", "WORKNET_API_KEY",
           "GCP_SERVICE_ACCOUNT_JSON", "GOOGLE_SHEET_URL", "WEBHOOK_URL"):
    if _ENV.get(_k) and not os.environ.get(_k):
        os.environ[_k] = _ENV[_k]
os.environ.setdefault("GOOGLE_SHEET_URL", _ENV.get("QRADAR_SHEET", ""))
os.environ.setdefault("GCP_SERVICE_ACCOUNT_JSON",
                      (HERE / "gcp-key.json").read_text(encoding="utf-8") if (HERE / "gcp-key.json").exists() else "")


def norm_law(s):
    return re.sub(r"\s+", " ", str(s or "").strip())


def norm_date(v):
    d = "".join(ch for ch in str(v or "") if ch.isdigit())
    return d[:8] if len(d) >= 8 else ""


def _loose(s):
    t = re.sub(r"[\s「」·ㆍ,]", "", str(s or ""))
    t = re.sub(r"\([^)]*\)$", "", t)
    return t


# ── 개명 매핑 (구명칭 느슨키 → 현행명 후보들, 순서대로 시도) ─────────────
RENAME_MAP = {
    "근로자직업능력개발법": ["국민 평생 직업능력 개발법"],
    "승강기시설안전관리법": ["승강기 안전관리법"],
    "소프트웨어산업진흥법": ["소프트웨어 진흥법"],
    "소재부품전문기업등의육성에관한특별조치법": [
        "소재·부품·장비산업 경쟁력 강화 및 공급망 안정화를 위한 특별조치법",
        "소재·부품·장비산업 경쟁력강화를 위한 특별조치법",
    ],
    "풍수해보험법": ["풍수해·지진재해보험법", "풍수해보험법"],
    "행정안전부소관비상대비자원관리법": ["비상대비에 관한 법률"],
    "비상대비자원관리법": ["비상대비에 관한 법률"],
    # ── 사용자 실조사 추가분 (2026-07): 미해결 3건의 열쇠 ──
    "수질및수생태계보전에관한법률": ["물환경보전법"],
    "화재예방소방시설설치유지및안전관리에관한법률": ["소방시설 설치 및 관리에 관한 법률"],
}
_SUFFIXES = ["시행규칙", "시행령"]  # 긴 것부터


def resolve_candidates(ledger_name):
    """대장 법령명 → 현행명 후보 리스트 (개명 매핑 + 접미어 보존 + 원명 폴백)"""
    ln = norm_law(ledger_name)
    ll = _loose(ln)
    suffix = ""
    base = ll
    for sfx in _SUFFIXES:
        if ll.endswith(sfx):
            suffix, base = sfx, ll[: -len(sfx)]
            break
    cands = []
    if base in RENAME_MAP:
        for new in RENAME_MAP[base]:
            cands.append(f"{new} {suffix}".strip() if suffix else new)
    cands.append(ln)  # 폴백: 원명 그대로 (A그룹 45건은 이걸로 잡힘)
    # 무공백 표기 대비: 공백 넣은 변형은 API substring이 어차피 못 잡으니 후보는 위로 충분
    seen, out = set(), []
    for c in cands:
        if c not in seen:
            seen.add(c); out.append(c)
    return out


# ── 법제처 조회 ──────────────────────────────────────────────────────
def _get(url_params):
    import requests
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        r = requests.get("https://www.law.go.kr/DRF/lawSearch.do", params=url_params,
                         headers=headers, timeout=30)
        if r.status_code == 200 and r.text.strip():
            return ET.fromstring(r.text)
    except Exception:
        return None
    return None


def find_current(api_key, candidates):
    """현행법(target=law)에서 후보명으로 검색 → 정확(느슨) 일치 노드의
    (현행명, MST, 시행일자, 제개정구분, 소관부처, 공포번호, 공포일자) 반환"""
    for cand in candidates:
        root = _get({"OC": api_key, "target": "law", "type": "XML",
                     "query": cand, "display": 100})
        if root is None:
            continue
        exact, contains = None, None
        for law in root.findall(".//law"):
            nm = norm_law(law.findtext("법령명한글", ""))
            if not nm:
                continue
            node = (nm,
                    law.findtext("법령일련번호", ""),
                    norm_date(law.findtext("시행일자", "")),
                    (law.findtext("제개정구분명", "") or "").strip(),
                    norm_law(law.findtext("소관부처명", "")),
                    re.sub(r"\D", "", law.findtext("공포번호", "") or ""),
                    norm_date(law.findtext("공포일자", "")))
            if _loose(nm) == _loose(cand):
                exact = node; break
            if contains is None and len(_loose(cand)) >= 6 and _loose(cand) in _loose(nm):
                contains = node
        hit = exact or contains
        if hit and hit[1]:
            return hit, cand
        time.sleep(0.2)
    return None, None


def fetch_body(api_key, mst, max_chars=60000):
    """lawService로 현행법 전문 수집 → 마크다운 (조문 전체 + 별표 일부, 길이 상한)"""
    import requests
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        r = requests.get("https://www.law.go.kr/DRF/lawService.do",
                         params={"OC": api_key, "target": "law", "MST": mst, "type": "XML"},
                         headers=headers, timeout=60)
        if r.status_code != 200 or not r.text.strip():
            return ""
        root = ET.fromstring(r.text)
    except Exception:
        return ""
    try:
        from hrdk_law_core.scraper import clean_to_markdown as _md
    except Exception:
        def _md(t, c):
            return f"### {t}\n{c}".strip()
    parts = []
    reason = ""
    for tag in (".//개정이유", ".//제개정이유"):
        n = root.find(tag)
        if n is not None and n.text:
            reason += n.text.strip() + "\n"
    if reason:
        parts.append("## [제·개정이유]\n" + reason[:3000])
    for jomun in root.findall(".//조문단위"):
        if jomun.attrib.get("조문여부") != "조문":
            continue
        t = jomun.find("조문제목"); c = jomun.find("조문내용")
        title = (t.text or "") if t is not None else ""
        content = (c.text or "") if c is not None else ""
        if title or content:
            parts.append(_md(title, content[:1500]))
        if sum(len(p) for p in parts) > max_chars:
            parts.append("(…이하 조문 생략: 길이 상한)")
            break
    stars = [s.text.strip() for s in root.findall(".//별표내용") if s is not None and s.text]
    if stars:
        parts.append("## [별표]\n" + "\n".join(stars)[:8000])
    return "\n\n".join(parts)


# ── 시트 ─────────────────────────────────────────────────────────────
def open_ws():
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    key = _ENV["QRADAR_SHEET"]
    m = re.search(r"/d/([A-Za-z0-9_-]+)", key)
    if m:
        key = m.group(1)
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        json.loads((HERE / "gcp-key.json").read_text(encoding="utf-8"), strict=False),
        ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"])
    return gspread.authorize(creds).open_by_key(key).worksheet(MAIN_TAB)


def scan(ws):
    values = ws.get_all_values()
    header = [h.strip() for h in values[0]]
    ix = {h: i for i, h in enumerate(header)}
    targets, key_rows, rowinfo = [], {}, {}
    for rn, row in enumerate(values[1:], start=2):
        def cell(h):
            i = ix.get(h, 10 ** 9)
            return str(row[i]).strip() if i < len(row) else ""
        law = cell("법령명")
        if not law:
            continue
        k = f"{norm_law(law)}|{norm_date(cell('시행일자'))}"
        key_rows.setdefault(k, []).append(rn)
        rowinfo[rn] = (cell("MST_ID"), law, norm_date(cell("시행일자")) or cell("시행일자"))
        if not cell("개정유형"):
            targets.append({"row": rn, "mst": cell("MST_ID"), "law": law,
                            "date": norm_date(cell("시행일자")) or cell("시행일자"),
                            "review": cell("검토사유")})
    all_keys = {k: v[0] for k, v in key_rows.items()}
    dups = {k: v for k, v in key_rows.items() if len(v) > 1}
    return header, ix, targets, all_keys, dups, rowinfo


def _preflight(need_llm):
    ok = True
    for k, msg in (("QRADAR_SHEET", "새 Q-RADAR 시트 URL"), ("LAW_API_KEY", "법제처 API 키")):
        if not _ENV.get(k):
            print(f"⚠️ .env에 {k}({msg})가 없습니다."); ok = False
    if need_llm and not _ENV.get("GEMINI_API_KEY"):
        print("⚠️ .env에 GEMINI_API_KEY가 없습니다 (재분석=AI 호출에 필요).")
        print("   👉 로컬 백필 때 쓰던 그 키를 한 줄 추가:  GEMINI_API_KEY=...")
        ok = False
    if not (HERE / "gcp-key.json").exists():
        print("🔑 gcp-key.json이 이 폴더에 없습니다."); ok = False
    try:
        import hrdk_law_core  # noqa
        try:
            from hrdk_law_core.certs import normalize_cert_string  # noqa — 최신 core 확인
        except Exception:
            print("⚠️ 설치된 hrdk-law-core가 구버전입니다 (필수 함수 없음).")
            print("   👉 업그레이드 한 줄:")
            print("      pip install --upgrade --force-reinstall git+https://github.com/gjtjdwns1008-dev/hrdk-law-core.git")
            ok = False
    except Exception:
        print("⚠️ hrdk-law-core가 이 PC에 설치돼 있지 않습니다.")
        print("   👉 터미널에서:  pip install git+https://github.com/gjtjdwns1008-dev/hrdk-law-core.git")
        ok = False
    if not ok:
        print("\n준비물을 채운 뒤 다시:  python reanalyze_ghosts.py")
    return ok


# ── 1) 미리보기 ───────────────────────────────────────────────────────
def preview():
    print("=" * 62); print("🔬 유령 재분석 — 미리보기 (대상 확정, AI 호출 없음)"); print("=" * 62)
    ws = open_ws()
    header, ix, targets, all_keys, dups, rowinfo = scan(ws)
    print(f"  📥 개정유형 빈 행: {len(targets)}행 → 현행법 매칭 시작")
    api = _ENV["LAW_API_KEY"]
    plans, conflicts, unresolved = [], [], []
    for i, t in enumerate(targets, 1):
        hit, used = find_current(api, resolve_candidates(t["law"]))
        if not hit:
            unresolved.append((t["row"], t["mst"], t["date"], t["law"], "현행법 검색 실패"))
        else:
            cur_name, mst_no, cur_date, gbn, ministry, pn, pd = hit
            renamed = _loose(cur_name) != _loose(t["law"])
            new_key = f"{cur_name}|{cur_date}"
            if new_key in all_keys and all_keys[new_key] != t["row"]:
                conflicts.append({"ghost_row": t["row"], "ghost_mst": t["mst"],
                                  "old_name": t["law"], "old_date": t["date"],
                                  "cur_name": cur_name, "cur_date": cur_date,
                                  "survivor_row": all_keys[new_key]})
            else:
                plans.append({"row": t["row"], "mst": t["mst"], "old_name": t["law"],
                              "old_date": t["date"], "cur_name": cur_name, "cur_date": cur_date,
                              "gbn": gbn, "mst_no": mst_no, "ministry": ministry,
                              "prom_num": pn, "prom_date": pd, "renamed": renamed})
        if i % 10 == 0 or i == len(targets):
            print(f"     …{i}/{len(targets)}")
        time.sleep(0.2)

    # ── 계획 내부 중복: 두 유령이 같은 현행키로 재분석 예약 → 뒤 것은 흡수로 전환 ──
    #    (같은 현행법으로 둘 다 갱신하면 대장에 쌍둥이 행이 생기는 사고 방지)
    _seen_new, _kept = {}, []
    for p in plans:
        nk = f"{p['cur_name']}|{p['cur_date']}"
        if nk in _seen_new:
            conflicts.append({"ghost_row": p["row"], "ghost_mst": p["mst"],
                              "old_name": p["old_name"], "old_date": p["old_date"],
                              "cur_name": p["cur_name"], "cur_date": p["cur_date"],
                              "survivor_row": _seen_new[nk], "kind": "dup"})
        else:
            _seen_new[nk] = p["row"]
            _kept.append(p)
    plans = _kept

    # ── 대장 전수 중복키: 이미 이중으로 존재하는 행(과거 사고 포함) → 뒤 행 흡수 ──
    _busy = {p["row"] for p in plans} | {c["ghost_row"] for c in conflicts}
    _ghosted = {c["ghost_row"] for c in conflicts}
    for _k, _rns in dups.items():
        _head = _rns[0]
        if _head in _ghosted:
            continue
        for _rn in _rns[1:]:
            if _rn in _busy:
                continue
            _mst, _law, _date = rowinfo.get(_rn, ("", "", ""))
            conflicts.append({"ghost_row": _rn, "ghost_mst": _mst, "old_name": _law,
                              "old_date": _date, "cur_name": _law, "cur_date": _date,
                              "survivor_row": _head, "kind": "dup"})
            _busy.add(_rn)

    renamed_n = sum(1 for p in plans if p["renamed"])
    print(f"  ✅ 재분석 확정 {len(plans)}행 (개명 {renamed_n}) / ★흡수예정 {len(conflicts)}행★ / 미해결 {len(unresolved)}")

    from openpyxl import Workbook
    wb = Workbook(); w1 = wb.active; w1.title = "재분석대상"
    w1.append(["행", "MST_ID", "구 법령명", "구 시행일자", "→ 현행명", "현행 시행일자", "제개정구분", "개명여부"])
    for p in plans:
        w1.append([p["row"], p["mst"], p["old_name"], p["old_date"], p["cur_name"],
                   p["cur_date"], p["gbn"], "개명" if p["renamed"] else ""])
    w2 = wb.create_sheet("흡수예정")
    w2.append(["유령 행", "유령 MST_ID", "구 법령명", "구 일자", "현행명", "현행일자",
               "생존 행", "처리"])
    for c in conflicts:
        w2.append([c["ghost_row"], c["ghost_mst"], c["old_name"], c["old_date"],
                   c["cur_name"], c["cur_date"], c["survivor_row"],
                   "생존행 검토사유에 흡수 태그 기록 후 유령행 삭제"])
    w3 = wb.create_sheet("미해결"); w3.append(["행", "MST_ID", "일자", "법령명", "사유"])
    for u in unresolved:
        w3.append(list(u))
    w4 = wb.create_sheet("통계")
    for k, v in [("대상", len(targets)), ("재분석 확정", len(plans)), ("이 중 개명", renamed_n),
                 ("흡수(유령 삭제 예정)", len(conflicts)), ("미해결", len(unresolved))]:
        w4.append([k, v])
    try:
        wb.save(PREVIEW_PATH); print(f"  💾 미리보기: {PREVIEW_PATH}")
    except PermissionError:
        alt = str(HERE / f"재분석_미리보기_{time.strftime('%H%M%S')}.xlsx")
        wb.save(alt); print(f"  💾 미리보기: {alt} (기존 파일이 열려 있음)")
    # 반영 단계용 계획 저장
    (HERE / "_reanalyze_plan.json").write_text(
        json.dumps({"reanalyze": plans, "absorb": conflicts}, ensure_ascii=False), encoding="utf-8")
    print("\n✅ 미리보기 완료. '재분석대상'·'흡수예정' 시트 확인 후 → 2번(반영)")
    print("   (2번은 이 계획 파일을 그대로 실행합니다 — 확인한 것과 동일 보장)")


# ── 2) 재분석 반영 ────────────────────────────────────────────────────
def apply():
    planp = HERE / "_reanalyze_plan.json"
    if not planp.exists():
        print("⛔ 먼저 1번(미리보기)을 실행해 계획을 만들어 주세요."); return
    _plan = json.loads(planp.read_text(encoding="utf-8"))
    if not isinstance(_plan, dict):
        print("⛔ 계획 파일이 구버전 형식입니다 (흡수 목록 없음).")
        print("   👉 먼저 1번(미리보기)을 다시 실행해 새 계획을 만들어 주세요.")
        return
    plans = _plan["reanalyze"]
    absorbs = _plan.get("absorb", [])
    print("=" * 62)
    print(f"🧪 유령 정비 반영 — 재분석 {len(plans)}행 + 흡수 {len(absorbs)}행")
    print("=" * 62)

    sys.path.insert(0, str(REPO))
    from brain import run_ai_analysis                     # 통합 brain 그대로
    from config import COLUMNS
    from hrdk_law_core.certs import get_qnet_certs_text, normalize_cert_string
    try:
        from hrdk_law_core.worknet import get_worknet_job_count
    except Exception:
        get_worknet_job_count = None
    try:
        sys.path.insert(0, str(REPO))
        from knowledge import QRadarKB
        from hrdk_law_core.hybrid import verify_with_krivet
        kb = QRadarKB(str(REPO / "hrdk_law.db"))
    except Exception:
        kb, verify_with_krivet = None, None

    from gspread.utils import rowcol_to_a1  # noqa (row range 문자열 직접 구성)
    ws = open_ws()
    certs_text = get_qnet_certs_text()
    api = _ENV["LAW_API_KEY"]
    wkey = _ENV.get("WORKNET_API_KEY", "")
    end_col = chr(ord("A") + len(COLUMNS) - 1) if len(COLUMNS) <= 26 else "X"

    ok_n, fail = 0, []
    for i, p in enumerate(plans, 1):
        print(f"  [{i}/{len(plans)}] {p['cur_name'][:30]} ({p['cur_date']})")
        body = fetch_body(api, p["mst_no"])
        if not body:
            fail.append((p["row"], p["cur_name"], "원문 수집 실패")); continue
        law = {"법령명": p["cur_name"], "시행일자": p["cur_date"], "소관부처": p["ministry"],
               "공포번호": p["prom_num"], "공포일자": p["prom_date"],
               "링크": f"https://www.law.go.kr/법령/{p['cur_name']}", "원본": body}
        ok, rel, info = run_ai_analysis(law, certs_text)
        if not ok:
            fail.append((p["row"], p["cur_name"], f"AI 실패: {str(info.get('error'))[:40]}")); continue
        # 사실값·이력·정돈
        info["MST_ID"] = p["mst"]
        info["개정유형"] = p["gbn"] or info.get("개정유형", "")
        std, dropped = normalize_cert_string(info.get("관련 종목", ""), year=2026)
        info["관련 종목"] = std
        tag = f"[기준선 재분석: 구 {p['old_name']}|{p['old_date']}]" if p["renamed"] \
              else f"[기준선 재분석: 구 시행일자 {p['old_date']}]"
        extra = f" / 사전외 종목 제외: {', '.join(dropped)}" if dropped else ""
        info["검토사유"] = (str(info.get("검토사유", "")) + " " + tag + extra).strip()
        if dropped:
            info["검토필요"] = "O"
        if str(info.get("우대여부", "")).strip() == "O":
            if get_worknet_job_count and wkey:
                info["워크넷 실시간 구인건수"] = get_worknet_job_count(info.get("관련 종목", ""), api_key=wkey)
            if kb and verify_with_krivet:
                try:
                    info = verify_with_krivet(info, kb)
                except Exception:
                    pass
        row_vals = [info.get(c, "") for c in COLUMNS]
        try:
            ws.update(range_name=f"A{p['row']}:{end_col}{p['row']}", values=[row_vals])
            ok_n += 1
        except Exception as e:
            fail.append((p["row"], p["cur_name"], f"시트 기록 실패: {str(e)[:40]}"))
        time.sleep(1.0)

    # ── 흡수 처리: ① 생존행 검토사유에 족보 태그 → ② 유령행 삭제(아래 행부터) ──
    #    순서가 생명: 재분석 갱신·태그 기록(행번호 기반)이 모두 끝난 뒤에만 삭제.
    if absorbs:
        print(f"\n  🫧 유령 흡수 {len(absorbs)}행 처리 중…")
        from collections import defaultdict
        vals = ws.get_all_values()
        hdr2 = [h.strip() for h in vals[0]]
        rv_i = hdr2.index("검토사유")

        def _col_letter(n):
            s = ""
            while n:
                n, r = divmod(n - 1, 26)
                s = chr(65 + r) + s
            return s

        tags = defaultdict(list)
        for a in absorbs:
            _label = "중복행 흡수" if a.get("kind") == "dup" else "기준선 유령 흡수"
            tags[a["survivor_row"]].append(
                f"[{_label}: {a['ghost_mst']}({a['old_name']}|{a['old_date']})]")
        updates = []
        for srow, tg in sorted(tags.items()):
            cur = vals[srow - 1][rv_i] if srow - 1 < len(vals) and rv_i < len(vals[srow - 1]) else ""
            merged = (str(cur).strip() + " " + " ".join(tg)).strip()
            updates.append({"range": f"{_col_letter(rv_i + 1)}{srow}", "values": [[merged]]})
        for i in range(0, len(updates), 400):
            ws.batch_update(updates[i:i + 400], value_input_option="RAW")
        print(f"     족보 태그 기록: 생존행 {len(tags)}곳")
        ghost_rows = sorted({a["ghost_row"] for a in absorbs}, reverse=True)
        for j, gr in enumerate(ghost_rows, 1):
            ws.delete_rows(gr)
            if j % 10 == 0 or j == len(ghost_rows):
                print(f"     유령행 삭제 {j}/{len(ghost_rows)} (아래부터 — 행번호 보존)")
            time.sleep(0.6)

    print(f"\n🎉 정비 완료: 재분석 성공 {ok_n} / 실패 {len(fail)} / 흡수 삭제 {len(absorbs)}행")
    for f in fail[:10]:
        print(f"   ⚠️ 행{f[0]} {f[1][:24]} — {f[2]}")
    if fail:
        print("   (실패분은 다시 1→2 실행하면 재시도됩니다 — 성공분은 개정유형이 채워져 자동 제외)")
    planp.unlink(missing_ok=True)
    print("\n※ SQLite는 손대지 않았습니다 — 다음 일일 실행이 시트→db 재구축으로 자동 흡수합니다.")


if __name__ == "__main__":
    args = [a.lower() for a in sys.argv[1:]]
    if "preview" in args:
        _preflight(False) and preview(); sys.exit(0)
    if "apply" in args:
        _preflight(True) and apply(); sys.exit(0)
    print("=" * 62)
    print("🔬 Q-RADAR 유령 재분석 도구 — 쉬운 모드")
    print("   (개명 매핑 + 현행법 기준 재분석 → 그 행 갱신)")
    print("=" * 62)
    print("\n무엇을 할까요?")
    print("  1) 미리보기 — 어떤 행이 어떤 현행법으로 재분석될지 목록만 (AI 호출 없음)")
    print("  2) 재분석 반영 — 1번 계획대로 원문수집→AI분석→행 갱신 (시간·비용 발생)")
    print("  0) 종료")
    choice = input("\n번호를 입력하고 Enter: ").strip()
    if choice == "1":
        _preflight(False) and preview()
    elif choice == "2":
        if not _preflight(True):
            sys.exit(0)
        confirm = input("⚠️ AI 재분석 후 해당 행들을 실제로 덮어씁니다. 진행하려면 '반영' 입력: ").strip()
        if confirm == "반영":
            apply()
        else:
            print("취소했습니다. (아무 변화 없음)")
    else:
        print("종료합니다. (아무 변화 없음)")
