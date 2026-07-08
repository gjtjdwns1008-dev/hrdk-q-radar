# -*- coding: utf-8 -*-
"""
reanalyze_flagged.py — 검토필요 정리 도구 (모드C 재산정 + 모드B 별표 심층 재분석)
=============================================================================
배경(2026-07-06): 구 기준(신뢰도 '보통'이면 O)이 검토필요를 남발해
1,255/1,638행이 O — HITL 변별력 상실. 별도로, 파일 전용 별표 탓에
"별표가 본문에 없다"는 사유 68행과 환각 오염 사유 1행이 존재.

두 모드의 분업 (순서: B 먼저 → C):
  [모드B] 별표류 사유 68행 — 법제처 재수집(annex 심층 수집 포함, 로컬 국내IP)
          → 통합 brain 풀 재분석(신 기준 프롬프트) → 분석 17칸(F:V) 교체.
          ※ 오염 사유 행(가축분뇨×철도)도 이 큐 소속 — 재분석이 자연 치유.
  [모드C] 나머지 검토필요 행 — 신 기준으로 O/X를 '순수 규칙' 재계산(LLM 0회):
          O 유지 = 신뢰도 '낮음' / 활용도 '대폭 증가·감소' / 중처법 애매값.
          탈락 행은 검토필요=X + 검토사유 비움. 신 기준 ⊂ 구 기준이라
          X→O 신규 발생은 구조적으로 없음.
  ※ C는 별표큐(B 대상)를 건드리지 않음 — B의 표적 보존.

안전핀: 미리보기 xlsx → 눈 확인 → 반영 / 반영 직전 실시간 재검증 /
        모드B 캐시 이어달리기(재과금 0) / C·B 모두 지정 칸만 기록.

실행: python reanalyze_flagged.py   (backfill_usage.py와 같은 폴더·.env 재사용)
준비물: .env(QRADAR_SHEET, GEMINI_API_KEY, LAW_API_KEY) + gcp-key.json
        + 로컬 core 최신 설치(재발방지판: pip install -e <repo>/core → pyhwp 포함)
"""
import os
import re
import sys
import json
import time
import shutil
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import backfill_usage as bu  # _retry / open_ws / _nsp / _col_letter / .env 로딩 재사용

# ★자급자족 번들(v1.2): 신판 core(annex 포함)+brain 동봉 — 로컬 설치본을 그림자 처리
_BUNDLE = HERE / "_bundle"
if _BUNDLE.exists():
    sys.path.insert(0, str(_BUNDLE))

REPO = HERE.parent
CACHE_B = HERE / "usage_cache_flagged.json"
PREVIEW_C = str(HERE / "재산정_모드C_미리보기.xlsx")
PREVIEW_B = str(HERE / "재분석_모드B_미리보기.xlsx")

TARGET_REL = ("연관높음", "단순관련")
BIG = ("대폭 증가", "대폭 감소")
MJC_OK = ("대상", "비대상", "")           # 이 밖의 값 = '애매' → O 유지
ANNEX_RE = re.compile(r"별표|서식|별지")

# 시트 F~V (17칸): 모드B가 교체하는 분석 칸 — 반영 시 헤더와 대조 검증
AN_COLS = ["연관도", "우대여부", "관련 종목", "주요 제·개정내용", "활용도_구분",
           "활용도_상세", "조문 요약", "우대분류", "Track1_취급유형", "Track1_위험도",
           "Track2_효용코드", "중처법대상", "상세 분석 결과", "근거조문",
           "AI신뢰도", "검토필요", "검토사유"]


# ── 공통: 스캔·분류 ──────────────────────────────────────────────────────
def classify(records):
    """행 분류 → (모드B 큐, C에서 O 유지, C에서 O→X 정리) 각 리스트."""
    qB, keep, clear = [], [], []
    for i, r in enumerate(records):
        rel = str(r.get("연관도", "") or "").strip()
        if rel not in TARGET_REL:
            continue
        cur_o = str(r.get("검토필요", "") or "").strip().upper() == "O"
        if not cur_o:
            continue  # 신 기준 ⊂ 구 기준 → X행은 손대지 않음
        why = str(r.get("검토사유", "") or "")
        item = {"row": i + 2, "mst": str(r.get("MST_ID", "")).strip(),
                "date": str(r.get("시행일자", "")).strip(), "rel": rel,
                "law": str(r.get("법령명", "")).strip(), "why": why.strip()}
        if ANNEX_RE.search(why):
            qB.append(item)
            continue
        trust = str(r.get("AI신뢰도", "") or "").strip()
        usage = str(r.get("활용도_구분", "") or "").strip()
        mjc = str(r.get("중처법대상", "") or "").strip()
        new_o = (trust == "낮음") or (usage in BIG) or (mjc not in MJC_OK)
        item["basis"] = ("신뢰도 낮음" if trust == "낮음" else
                         f"활용도 {usage}" if usage in BIG else
                         f"중처법 애매값('{mjc}')" if mjc not in MJC_OK else "")
        (keep if new_o else clear).append(item)
    return qB, keep, clear


def _dup_mst(records):
    """중복 MST 전수 탐지 → {mst: [행번호,...]} (v1.3 무결성 게이트)"""
    seen, dups = {}, {}
    for i, r in enumerate(records):
        m = str(r.get("MST_ID", "")).strip()
        if not m:
            continue
        if m in seen:
            dups.setdefault(m, [seen[m]]).append(i + 2)
        else:
            seen[m] = i + 2
    return dups


def _gate(records, ctx):
    """쓰기 작업 공통 게이트 — 중복 MST 발견 시 행번호 보고 후 중단 신호(True)."""
    dups = _dup_mst(records)
    if dups:
        print(f"🚫 [{ctx}] 중복 MST {len(dups)}종 발견 — 안전을 위해 반영을 중단합니다.")
        for m, rows in list(dups.items())[:6]:
            print(f"   · {m} → 시트 행 {rows}")
        print("   👉 중복 행을 정리(한 행 삭제)한 뒤 다시 실행하세요.")
        return True
    return False


# ── 모드C: 미리보기 / 반영 ──────────────────────────────────────────────
def write_preview_c(qB, keep, clear):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "모드C_O→X 정리 대상"
    ws.append(["MST_ID", "시행일자", "연관도", "법령명", "기존 검토사유(삭제 예정)"])
    for t in clear:
        ws.append([t["mst"], t["date"], t["rel"], t["law"], t["why"][:150]])
    w2 = wb.create_sheet("모드C_O 유지")
    w2.append(["MST_ID", "법령명", "유지 근거"])
    for t in keep:
        w2.append([t["mst"], t["law"], t["basis"]])
    w3 = wb.create_sheet("모드B 큐(별표류, C 무접촉)")
    w3.append(["MST_ID", "시행일자", "법령명", "검토사유"])
    for t in qB:
        w3.append([t["mst"], t["date"], t["law"], t["why"][:150]])
    wb.save(PREVIEW_C)


def apply_c(ws):
    """반영 직전 실시간 재계산 → O→X 행의 검토필요·검토사유 2칸만 정리."""
    header = bu._retry("헤더 읽기", lambda: ws.row_values(1))
    uL = bu._col_letter(header.index("검토필요") + 1)
    vL = bu._col_letter(header.index("검토사유") + 1)
    records = bu._retry("대장 읽기", lambda: ws.get_all_records())
    if _gate(records, "모드C"):
        return None
    qB, keep, clear = classify(records)
    updates = []
    for t in clear:
        updates.append({"range": f"{uL}{t['row']}", "values": [["X"]]})
        updates.append({"range": f"{vL}{t['row']}", "values": [[""]]})
    done = 0
    for i in range(0, len(updates), 20):
        chunk = updates[i:i + 20]
        bu._retry(f"기록 {i + 1}~{i + len(chunk)}", lambda c=chunk: ws.batch_update(c))
        done += len(chunk)
        print(f"  ✍️ {done}/{len(updates)}칸")
    return len(clear), len(keep), len(qB)


# ── 모드B: 재수집(별표 심층 포함) + brain 풀 재분석 ─────────────────────
def _certs_text():
    from hrdk_law_core.certs import get_qnet_certs_text
    return get_qnet_certs_text()


def fetch_pinpoint(law_name):
    """★v1.4 핀포인트: 법령명 → 현행 MST → 그 법령 1건만 본문+별표 수집.
    (날짜 전체 스캔 폐지 — 재분석 속도 수십 배)  반환: (law|None, note, 진단census)"""
    try:
        import xml.etree.ElementTree as ET
        import requests as _rq
        from reanalyze_ghosts import find_current, fetch_body
        try:
            from reanalyze_ghosts import resolve_candidates as _cands
        except ImportError:
            _cands = lambda nm: [nm]
        from hrdk_law_core.annex import build_annex_sections, census
    except Exception as e:
        return None, f"모듈 로딩 실패({str(e)[:30]})", ""
    key = bu._ENV.get("LAW_API_KEY", "") or os.environ.get("LAW_API_KEY", "")
    if not key:
        return None, "LAW_API_KEY 없음", ""
    hit, used = find_current(key, _cands(law_name))
    if not hit:
        return None, "현행 검색 미발견", ""
    nm, mst, ef, gubun, ministry, pn, pd_ = hit
    body = fetch_body(key, mst)
    if not body:
        return None, "현행 본문 수집 실패", ""
    cen = ""
    try:
        H = {"User-Agent": "Mozilla/5.0"}
        r = _rq.get("https://www.law.go.kr/DRF/lawService.do", headers=H, timeout=60,
                    params={"OC": key, "target": "law", "MST": mst, "type": "XML"})
        root = ET.fromstring(r.text)
        cen = census(root)
        _g = lambda u: _rq.get(u, timeout=30, headers=H).content
        _gt = lambda u: _rq.get(u, timeout=30, headers=H).text
        ax_t, ax_s = build_annex_sections(root, _g, law_name=nm, api_key=key,
                                          http_get_text=_gt)
        if ax_t:
            body += "\n\n" + ax_t
        if ax_s:
            body += "\n\n" + ax_s
    except Exception:
        pass
    law = {"법령명": nm, "시행일자": ef, "소관부처": ministry, "공포번호": pn,
           "공포일자": pd_, "원본": body, "링크": ""}
    return law, f"현행본(시행 {ef})", cen


def analyze(law, certs_text):
    from brain import run_ai_analysis
    return run_ai_analysis(law, certs_text)


def _save_b(c): CACHE_B.write_text(json.dumps(c, ensure_ascii=False, indent=1), encoding="utf-8")
def _load_b():
    if CACHE_B.exists():
        try: return json.loads(CACHE_B.read_text(encoding="utf-8"))
        except Exception: pass
    return {}


def run_b(targets, fetch_one=None, analyze_fn=None, cache=None, sleep=0.4):
    fetch_one = fetch_one or fetch_pinpoint
    analyze_fn = analyze_fn or analyze
    cache = cache if cache is not None else {}
    done = lambda v: bool(v.get("info")) and not v.get("err")
    pending = [t for t in targets if not done(cache.get(t["mst"], {}))]
    if len(targets) - len(pending):
        print(f"  ↻ 캐시 재사용 {len(targets) - len(pending)}건")
    print(f"  🎯 핀포인트 재분석 대상 {len(pending)}건 (법령별 직행 조회 — 날짜 스캔 없음)")
    certs_text = _certs_text() if pending else ""
    diag_path = HERE / "annex_진단.txt"
    diag_done, star = False, 0

    for k, t in enumerate(pending, 1):
        law, note, cen = fetch_one(t["law"])
        if not diag_done and cen:
            diag_path.write_text(f"대상 법령: {t['law']}\n{cen}\n", encoding="utf-8")
            diag_done = True
        if not law:
            cache[t["mst"]] = {"err": f"미발견({note})", "law": t["law"], "date": t["date"]}
            print(f"  [{k}/{len(pending)}] {t['mst']} {t['law'][:20]} → ❌ {note}")
            _save_b(cache); continue
        annexed = "별표(파일 추출" in law.get("원본", "")
        star += 1 if annexed else 0
        try:
            ok, rel, info = analyze_fn(law, certs_text)
        except Exception as e:
            ok, rel, info = False, "", {}
            print(f"    ⚠️ 분석 예외: {str(e)[:60]}")
        if not ok or not isinstance(info, dict):
            cache[t["mst"]] = {"err": "분석 실패", "law": t["law"], "date": t["date"]}
            print(f"  [{k}/{len(pending)}] {t['mst']} {t['law'][:20]} → ❌ 분석 실패")
        else:
            info = dict(info); info.setdefault("연관도", rel)
            try:  # ★분류체계 규약: 코드(한글 부연)
                from hrdk_law_core.certs import (label_track1_type,
                                                 label_track1_risk, label_track2_code)
                info["Track1_취급유형"] = label_track1_type(info.get("Track1_취급유형", ""))
                info["Track1_위험도"] = label_track1_risk(info.get("Track1_위험도", ""))
                info["Track2_효용코드"] = label_track2_code(info.get("Track2_효용코드", ""))
            except Exception as _le:
                print(f"    ⚠️ 라벨 주석화 생략: {str(_le)[:40]}")
            cache[t["mst"]] = {"err": "", "law": t["law"], "date": t["date"],
                               "annex": annexed, "old_rel": t["rel"], "fb": note,
                               "info": {c: str(info.get(c, "")) for c in AN_COLS}}
            flag = "⭐별표추출" if annexed else "별표없음/미확보"
            print(f"  [{k}/{len(pending)}] {t['mst']} {t['law'][:20]} → {info.get('연관도','')} "
                  f"[신뢰도 {info.get('AI신뢰도','')}] [{flag}] ({note})")
        if k % 3 == 0:
            _save_b(cache)
        time.sleep(sleep)
    _save_b(cache)
    if pending and star == 0:
        print(f"\n  🚨 별표 추출 0건 — {diag_path.name} 파일이 생성됐습니다. 이 파일을 공유해 주세요.")
    elif pending:
        print(f"\n  ⭐ 별표 추출 성공 {star}/{len(pending)}건")
    return cache


def write_preview_b(targets, cache):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "모드B 재분석 diff"
    ws.append(["MST_ID", "법령명", "별표추출", "연관도(구→신)", "신뢰도(신)",
               "검토필요(신)", "새 검토사유", "관련 종목(신, 앞 80자)", "비고"])
    ok = fail = 0
    for t in targets:
        v = cache.get(t["mst"], {})
        if v.get("err") or not v.get("info"):
            ws.append([t["mst"], t["law"], "", "", "", "", "", "", v.get("err", "미실행")]); fail += 1
            continue
        i = v["info"]; ok += 1
        ws.append([t["mst"], t["law"], "O" if v.get("annex") else "",
                   f"{v.get('old_rel','')}→{i['연관도']}", i["AI신뢰도"], i["검토필요"],
                   i["검토사유"][:120], i["관련 종목"][:80], v.get("fb", "")])
    wb.save(PREVIEW_B)
    return ok, fail


def apply_b(ws, cache):
    header = bu._retry("헤더 읽기", lambda: ws.row_values(1))
    f_idx = header.index(AN_COLS[0]) + 1
    seg = header[f_idx - 1: f_idx - 1 + len(AN_COLS)]
    if seg != AN_COLS:
        print("❌ 시트 열 배열이 예상(F:V 17칸)과 다릅니다 — 중단:", seg); sys.exit(1)
    fL, vL = bu._col_letter(f_idx), bu._col_letter(f_idx + len(AN_COLS) - 1)

    records = bu._retry("대장 읽기", lambda: ws.get_all_records())
    if _gate(records, "모드B"):
        return None
    by_mst = {str(r.get("MST_ID", "")).strip(): (i + 2, r) for i, r in enumerate(records)
              if str(r.get("MST_ID", "")).strip()}
    updates, skips = [], []
    for mst, v in cache.items():
        if v.get("err") or not v.get("info"):
            skips.append((mst, v.get("err") or "결과 없음")); continue
        if mst not in by_mst:
            skips.append((mst, "시트에 MST 없음")); continue
        row, r = by_mst[mst]
        if bu._nsp(r.get("법령명", "")) != bu._nsp(v.get("law", "")):
            skips.append((mst, "법령명 불일치(행 이동 의심) — 보존")); continue
        vals = [v["info"][c] for c in AN_COLS]
        updates.append({"range": f"{fL}{row}:{vL}{row}", "values": [vals]})
    done = 0
    for i in range(0, len(updates), 10):
        chunk = updates[i:i + 10]
        bu._retry(f"기록 {i + 1}~{i + len(chunk)}행", lambda c=chunk: ws.batch_update(c))
        done += len(chunk)
        print(f"  ✍️ {done}/{len(updates)}행 (F:V 17칸)")
    return done, skips


# ── 사고 복구: 백업 xlsx → B큐 행 F:V 원복 ────────────────────────────
def load_backup_records(xlsx_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["국가기술자격 관련법령"]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]
    return [dict(zip(header, [("" if v is None else str(v)) for v in r])) for r in rows]


def restore_b_rows(ws, backup_records):
    """[5] 오적용 복구 — 백업 기준 B큐(별표류) 행의 분석 17칸을 백업 값으로 되돌림.
    적용 안 된 행은 동일값 재기록(무해). 법령명 불일치 행은 보존."""
    qB, _, _ = classify(backup_records)
    if not qB:
        print("백업에서 B큐를 찾지 못했습니다."); return 0, []
    header = bu._retry("헤더 읽기", lambda: ws.row_values(1))
    f_idx = header.index(AN_COLS[0]) + 1
    if header[f_idx - 1: f_idx - 1 + len(AN_COLS)] != AN_COLS:
        print("❌ 열 배열 불일치 — 중단"); sys.exit(1)
    fL, vL = bu._col_letter(f_idx), bu._col_letter(f_idx + len(AN_COLS) - 1)
    bk_by = {t["mst"]: backup_records[t["row"] - 2] for t in qB}
    live = bu._retry("대장 읽기", lambda: ws.get_all_records())
    if _gate(live, "복구"):
        return None
    by_mst = {str(r.get("MST_ID", "")).strip(): (i + 2, r) for i, r in enumerate(live)
              if str(r.get("MST_ID", "")).strip()}
    updates, skips = [], []
    for mst, br in bk_by.items():
        if mst not in by_mst:
            skips.append((mst, "현 시트에 없음")); continue
        row, lr = by_mst[mst]
        if bu._nsp(lr.get("법령명", "")) != bu._nsp(br.get("법령명", "")):
            skips.append((mst, "법령명 불일치 — 보존")); continue
        vals = [str(br.get(c, "") or "") for c in AN_COLS]
        updates.append({"range": f"{fL}{row}:{vL}{row}", "values": [vals]})
    done = 0
    for i in range(0, len(updates), 10):
        chunk = updates[i:i + 10]
        bu._retry(f"복구 {i + 1}~{i + len(chunk)}행", lambda c=chunk: ws.batch_update(c))
        done += len(chunk); print(f"  ⏪ {done}/{len(updates)}행 원복")
    return done, skips


# ── 메뉴 ─────────────────────────────────────────────────────────────────
def _self_check():
    out = {}
    try:
        import hrdk_law_core.annex as _ax
        out["annex"] = _ax.__file__
    except ImportError:
        out["annex"] = ""
    try:
        import brain as _br
        out["brain"] = _br.__file__
    except Exception:
        out["brain"] = ""
    return out


def main():
    chk = _self_check()
    if chk["annex"]:
        src = "번들" if "_bundle" in chk["annex"] else "설치본"
        print(f"🔎 core: 신판(annex) ✓ [{src}] {chk['annex']}")
    else:
        print("🚨 core: annex 없음 — _bundle 폴더가 함께 압축해제됐는지 확인하세요.")
    if chk["brain"]:
        src = "번들" if "_bundle" in chk["brain"] else "레포"
        print(f"🔎 brain: [{src}] {chk['brain']}")
    if CACHE_B.exists():
        try:
            _n = len(json.loads(CACHE_B.read_text(encoding="utf-8")))
            print(f"🗂️ 모드B 캐시 {_n}건 존재 — 전면 재실행이라면 [6]으로 먼저 비우세요.")
        except Exception:
            pass
    try:
        from hrdk_law_core.annex import resolve_hwp5txt
        mode = resolve_hwp5txt()
        if mode[0] == "cmd":
            print(f"🔎 hwp5txt: ✓ 실행파일 [{mode[1][0]}]")
        elif mode[0] == "runpy":
            print("🔎 hwp5txt: ✓ 파이썬 내부 구동 모드 (PATH 무관)")
        else:
            print(f"🚨 hwp5txt: 사용 불가 — {mode[1]}")
            print("   해결: 이 파이썬에서  pip install pyhwp  후 재실행\n")
    except Exception as _he:
        print(f"⚠️ 별표 추출기 점검 실패: {str(_he)[:50]}")
    print("=" * 64)
    print("  검토필요 정리 도구 — 모드B(별표 심층 재분석) → 모드C(재산정)")
    print("=" * 64)
    while True:
        print("\n[1] 스캔·분류 리포트 (무LLM)")
        print("[2] 모드C 미리보기 생성 (O→X 정리안, 시트 무접촉)")
        print("[3] 모드C 반영 (검토필요·검토사유 2칸만)")
        print("[4] 모드B 재분석 + 미리보기 (수집+AI, 캐시 이어달리기)")
        print("[5] 모드B 반영 (분석 17칸 F:V 교체)")
        print("[6] 모드B 캐시 초기화 (재실행 전 필수)")
        print("[7] 🚑 사고 복구 — 백업 xlsx로 B큐 행(F:V) 원복  ※ 오적용 시 [7]→[6]→[4] 순서")
        print("[8] 🔄 현재 B큐만 캐시 무효화 — annex·프롬프트 업그레이드 후 부분 재분석 준비")
        print("[0] 종료")
        sel = input("번호 선택: ").strip()

        if sel == "7":
            path = input("백업 xlsx 경로 (예: C:\\...\\HRDK-Q-RADAR (1).xlsx): ").strip().strip('"')
            if not path or not Path(path).exists():
                print("❌ 파일을 찾을 수 없습니다."); continue
            recs = load_backup_records(path)
            qB, _, _ = classify(recs)
            print(f"백업 기준 B큐 {len(qB)}행의 분석 17칸(F:V)을 백업 값으로 되돌립니다.")
            if input("진행할까요? (y/n): ").strip().lower() != "y":
                continue
            ws = bu.open_ws()
            res = restore_b_rows(ws, recs)
            if res is None:
                continue
            done, skips = res
            print(f"\n✅ 원복 {done}행")
            for m, w in skips[:8]:
                print(f"   ⏭️ {m}: {w}")
            continue
        if sel == "8":
            ws = bu.open_ws()
            records = bu._retry("대장 읽기", lambda: ws.get_all_records())
            qB, _, _ = classify(records)
            cache = _load_b()
            hit = [t["mst"] for t in qB if t["mst"] in cache]
            print(f"현재 B큐 {len(qB)}건 중 캐시 보유 {len(hit)}건을 무효화합니다 (나머지 캐시 보존, 시트 무접촉).")
            if input("진행할까요? (y/n): ").strip().lower() != "y":
                continue
            for m in hit:
                del cache[m]
            _save_b(cache)
            print(f"✅ {len(hit)}건 무효화 — 이제 [4]가 이 행들만 신형 annex·프롬프트로 재분석합니다.")
            continue
        if sel == "6":
            if CACHE_B.exists() and input("캐시를 지우고 처음부터 재분석하시겠어요? (y/n): ").strip().lower() == "y":
                CACHE_B.unlink(); print("🧹 캐시 삭제 완료 — [4]를 다시 실행하세요.")
            elif not CACHE_B.exists():
                print("캐시가 이미 없습니다.")
            continue
        if sel == "0":
            return
        if sel in ("1", "2", "4"):
            ws = bu.open_ws()
            records = bu._retry("대장 읽기", lambda: ws.get_all_records())
            qB, keep, clear = classify(records)
            print(f"\n📋 검토필요 O 분류: 모드B 큐(별표류) {len(qB)} / C-유지 {len(keep)} / C-정리 {len(clear)}")
            if sel == "1":
                for t in qB[:5]:
                    print(f"   [B] {t['mst']} {t['law'][:28]}")
                dups = _dup_mst(records)
                print(f"🧬 무결성: 중복 MST {len(dups)}종" + (" ✓" if not dups else " 🚨"))
                for m, rows in list(dups.items())[:6]:
                    print(f"   · {m} → 행 {rows}")
                from collections import Counter
                nk = Counter()
                for r in records:
                    m = str(r.get("MST_ID", "")).strip()
                    if m:
                        nk[(bu._nsp(r.get("법령명", "")), str(r.get("시행일자", "")).strip())] += 1
                twins = sum(1 for v in nk.values() if v > 1)
                if twins:
                    print(f"ℹ️ 자연키 쌍둥이(같은 법령·시행일자, 다른 MST): {twins}종 — 추후 병합 과제")
                continue
            if sel == "2":
                write_preview_c(qB, keep, clear)
                print(f"📄 미리보기: {PREVIEW_C}\n   → 확인 후 [3]으로 반영")
                continue
            # sel == "4"
            if not qB:
                print("✅ 모드B 대상 없음"); continue
            if input(f"{len(qB)}건 재수집+재분석 시작? (y/n): ").strip().lower() != "y":
                continue
            cache = run_b(qB, cache=_load_b())
            ok, fail = write_preview_b(qB, cache)
            print(f"\n✅ 재분석 {ok} / 실패 {fail}")
            print(f"📄 diff 미리보기: {PREVIEW_B}\n   → 연관도·사유 diff 확인 후 [5]로 반영")

        elif sel == "3":
            print("※ 권장 순서: 모드B([4][5]) 완료 후 C — 별표큐는 C가 건드리지 않으니 순서 바뀌어도 안전합니다.")
            if input("모드C 반영(O→X 정리)을 진행할까요? (y/n): ").strip().lower() != "y":
                continue
            ws = bu.open_ws()
            res = apply_c(ws)
            if res is None:
                continue
            n_clear, n_keep, n_b = res
            print(f"\n✅ 정리 {n_clear}행 O→X / 유지 {n_keep} / 별표큐(무접촉) {n_b}")

        elif sel == "5":
            cache = _load_b()
            valid = sum(1 for v in cache.values() if v.get("info") and not v.get("err"))
            if not valid:
                print("⚠️ 모드B 캐시 없음 — 먼저 [4]"); continue
            print(f"캐시 유효 {valid}건. diff 미리보기를 확인하셨나요?")
            if input("F:V 17칸을 교체 기록할까요? (y/n): ").strip().lower() != "y":
                continue
            ws = bu.open_ws()
            res = apply_b(ws, cache)
            if res is None:
                continue
            done, skips = res
            print(f"\n✅ {done}행 반영")
            for m, why in skips[:8]:
                print(f"   ⏭️ {m}: {why}")
        else:
            print("0~5 중 선택")


if __name__ == "__main__":
    main()
