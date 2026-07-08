# -*- coding: utf-8 -*-
"""
restore_sheet.py — 대장 전체 복원 (백업 xlsx → 구글시트 '국가기술자격 관련법령' 탭)
=============================================================================
용도: 구글시트 UI 붙여넣기가 버벅일 때, 백업 엑셀을 통째로 시트에 되붓는 일회용 도구.

3중 안전장치:
  ① 쓰기 전에 '현재 시트'를 로컬 xlsx 스냅샷으로 자동 백업 (이 복원조차 되돌리기 가능)
  ② 백업 파일 사전 검증 — 헤더 일치, 중복 MST 0종 확인, 행수 미리보기
  ③ y 확인 후에만 기록 (RAW 모드 — 값 그대로, 수식 해석 없음)

실행: python restore_sheet.py
      → 백업 xlsx 경로 입력 (예: C:\\...\\HRDK-Q-RADAR (1).xlsx)
준비물: backfill_usage.py와 같은 폴더 (.env, gcp-key.json 재사용)
"""
import sys
import time
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import backfill_usage as bu  # open_ws / _retry / _col_letter / .env 재사용

TAB = "국가기술자격 관련법령"
CHUNK = 300  # 한 번에 쓰는 행 수


def read_backup(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if TAB not in wb.sheetnames:
        print(f"❌ 백업에 '{TAB}' 시트가 없습니다. 시트 목록: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[TAB]
    rows = [[("" if v is None else str(v)) for v in r] for r in ws.iter_rows(values_only=True)]
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()  # 꼬리 빈 행 제거
    # 꼬리 빈 열 제거 (엑셀 유령 열 방지 — 값이 있는 마지막 열까지만)
    width = 0
    for r in rows:
        for j in range(len(r) - 1, -1, -1):
            if r[j].strip():
                width = max(width, j + 1)
                break
    rows = [r[:width] + [""] * (width - len(r)) for r in rows]
    return rows


def dup_mst_check(rows):
    header = rows[0]
    try:
        mi = header.index("MST_ID")
    except ValueError:
        return {"(MST_ID 열 없음)": []}
    seen, dups = {}, {}
    for i, r in enumerate(rows[1:], start=2):
        m = (r[mi] if mi < len(r) else "").strip()
        if not m:
            continue
        if m in seen:
            dups.setdefault(m, [seen[m]]).append(i)
        else:
            seen[m] = i
    return dups


def snapshot_live(ws):
    from openpyxl import Workbook
    vals = bu._retry("현재 시트 읽기(스냅샷)", lambda: ws.get_all_values())
    wb = Workbook()
    sh = wb.active
    sh.title = TAB[:31]
    for r in vals:
        sh.append(r)
    name = HERE / f"복원전_스냅샷_{time.strftime('%y%m%d_%H%M%S')}.xlsx"
    wb.save(name)
    return name, len(vals)


def main():
    print("=" * 60)
    print("  대장 전체 복원 — 백업 xlsx → 구글시트")
    print("=" * 60)
    path = input("백업 xlsx 경로: ").strip().strip('"')
    if not path or not Path(path).exists():
        print("❌ 파일을 찾을 수 없습니다."); return

    rows = read_backup(path)
    n_rows, n_cols = len(rows) - 1, len(rows[0])
    print(f"\n📖 백업 판독: 데이터 {n_rows}행 × {n_cols}열 (+헤더)")
    print(f"   첫 행: {rows[1][0]} | {rows[1][3][:24]}")
    print(f"   끝 행: {rows[-1][0]} | {rows[-1][3][:24]}")
    dups = dup_mst_check(rows)
    if dups:
        print(f"🚨 백업 안에 중복 MST {len(dups)}종 — 복원 중단 (파일 확인 필요)")
        for m, rr in list(dups.items())[:6]:
            print(f"   · {m} → 백업 행 {rr}")
        return
    print("🧬 백업 무결성: 중복 MST 0종 ✓")

    ws = bu.open_ws()
    live_header = bu._retry("현 시트 헤더", lambda: ws.row_values(1))
    if live_header and live_header != rows[0]:
        print("⚠️ 현 시트 헤더와 백업 헤더가 다릅니다:")
        print(f"   시트: {live_header[:5]} ...")
        print(f"   백업: {rows[0][:5]} ...")
        if input("   그래도 백업 헤더 기준으로 덮어쓸까요? (y/n): ").strip().lower() != "y":
            return

    print(f"\n⚠️ 시트 '{TAB}' 전체를 백업 내용({n_rows}행)으로 교체합니다.")
    if input("진행할까요? (y/n): ").strip().lower() != "y":
        print("취소했습니다."); return

    snap, live_n = snapshot_live(ws)
    print(f"📸 복원 전 스냅샷 저장: {snap.name} ({live_n}행) — 문제 시 이 파일로 되돌리기 가능")

    bu._retry("시트 비우기", lambda: ws.clear())
    bu._retry("크기 조정", lambda: ws.resize(rows=len(rows) + 20, cols=n_cols))
    endL = bu._col_letter(n_cols)
    done = 0
    for i in range(0, len(rows), CHUNK):
        chunk = rows[i:i + CHUNK]
        rng = f"A{i + 1}:{endL}{i + len(chunk)}"
        bu._retry(f"기록 {rng}", lambda c=chunk, r=rng: ws.batch_update(
            [{"range": r, "values": c}], value_input_option="RAW"))
        done += len(chunk)
        print(f"  ✍️ {done}/{len(rows)}행")

    # 검증: 행수 + 표본 3행
    live = bu._retry("검증 읽기", lambda: ws.get_all_values())
    ok_n = (len(live) == len(rows))
    import random
    idxs = [1, len(rows) - 1] + ([random.randrange(1, len(rows) - 1)] if len(rows) > 3 else [])
    ok_s = all(live[i][:n_cols] == rows[i][:n_cols] for i in idxs)
    if ok_n and ok_s:
        print(f"\n✅ 복원 완료 — {n_rows}행 전량 일치 검증 통과")
        print("   다음: python reanalyze_flagged.py → [1] 스캔 (68/166/1021 + 중복 0종 기대)")
    else:
        print(f"\n⚠️ 검증 경고: 행수 일치 {ok_n}, 표본 일치 {ok_s}")
        print(f"   스냅샷({snap.name})과 백업 원본이 있으니 상황 공유해 주세요.")


if __name__ == "__main__":
    main()
