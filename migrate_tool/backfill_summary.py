# -*- coding: utf-8 -*-
"""
backfill_summary.py — 총괄현황표 이력 백필 (monitor → Q-RADAR)
=============================================================================
왜 필요한가?
  대장(분석 결과)은 이관했지만 총괄현황표(관제 이력: 일자별 총 검토건수)는
  monitor 시트에 남아 있었습니다. 그 결과 이관 이전 달(예: 2026-06)의 브리핑
  KPI '전수 검토 N건'이 비어 나옵니다. 이 도구가 monitor 총괄의 과거 행을
  Q-RADAR 총괄로 옮겨 '단일 진실 원천'을 완성합니다.

핵심 원칙 (v2): 관련 건수는 '대장의 파생값'
  monitor 총괄의 연관/단순 카운트는 흡수·재분석·중복정리 이전의 낡은 숫자입니다.
  → 총 검토건수만 monitor에서 가져오고(대장엔 없는 '그날 전수 우주의 크기'),
    연관높음·단순관련·우대건수는 통합 대장에서 그 날짜 행을 직접 세어 기록합니다.
  → 보너스: monitor에 없던 우대건수까지 채워집니다.

처리 방식 (upsert):
  · 없는 날짜 → 🔵 신규 삽입
  · 있는데 건수(B~E)가 전부 빈 행(예: 7/2 이음새) → 그 자리에 건수만 채워 소생
  · 건수가 이미 있는 행(예: 7/3 실전 기록) → 보존(스킵)

안전 설계 (관제 로직 무풍 보장 — 코드로 검증된 전제):
  · 이관 행의 모니터링 상태 = "🔵 이관" → read_last_success_date는 🟢/정상만
    보므로 크론의 밀린 날짜 계산에 절대 안 잡힘
  · sum_total_reviewed는 상태를 안 보고 월+건수만 합산 → 브리핑 KPI엔 잡힘
  · 이미 있는 날짜(A열 표시값 기준)는 건너뜀 → 몇 번을 돌려도 안전(멱등)

실행:  python backfill_summary.py  → 메뉴 (1 미리보기 / 2 반영 / 0 종료)
준비물(.env): QRADAR_SHEET(새 시트) + MONITOR_SHEET(구 monitor 시트) + gcp-key.json
"""
import re
import sys
import json
import time
from pathlib import Path

HERE = Path(__file__).resolve().parent
TAB = "총괄현황표"
PREVIEW_PATH = str(HERE / "총괄백필_미리보기.xlsx")
STAMP = "🔵 이관"
LOG_MSG = "monitor 총괄 이력 이관(백필 도구)"


def _load_env_file():
    env = {}
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


_ENV = _load_env_file()


def norm_date(v):
    d = "".join(ch for ch in str(v or "") if ch.isdigit())
    return d[:8] if len(d) >= 8 else ""


def _retry(what, fn, tries=3, wait=4):
    """일시적 통신 단절(보안SW·프록시의 순간 컷 등) 자동 재시도 — WinError 10053 대응."""
    for i in range(1, tries + 1):
        try:
            return fn()
        except Exception as e:
            if i == tries:
                print(f"  ⛔ {what} — {tries}회 모두 실패. 네트워크/VPN 확인 후 다시 실행하세요 (멱등: 그냥 재실행하면 이어서 진행).")
                raise
            print(f"  🔁 {what} 순간 끊김({type(e).__name__}) — {wait}초 후 재시도 {i}/{tries - 1}…")
            time.sleep(wait)


def _preflight():
    ok = True
    if not (HERE / "gcp-key.json").exists():
        print("🔑 gcp-key.json이 이 폴더에 없습니다."); ok = False
    for k, desc in (("QRADAR_SHEET", "새 Q-RADAR 시트 URL"),
                    ("MONITOR_SHEET", "구 law-monitor 시트 URL (이번에 한 줄 추가)")):
        if not _ENV.get(k):
            print(f"⚠️ .env에 {k}({desc})가 없습니다."); ok = False
    if not ok:
        print("\n준비물을 채운 뒤 다시:  python backfill_summary.py")
    return ok


MAIN_TAB = "국가기술자격 관련법령"


def _open_ss(url_key):
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    key = _ENV[url_key]
    m = re.search(r"/d/([A-Za-z0-9_-]+)", key)
    if m:
        key = m.group(1)
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        json.loads((HERE / "gcp-key.json").read_text(encoding="utf-8"), strict=False),
        ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"])
    return _retry(f"구글 시트 접속({url_key})",
                  lambda: gspread.authorize(creds).open_by_key(key))


def count_from_ledger(qss):
    """통합 대장에서 날짜별 (연관높음, 단순관련, 우대O) 직접 카운트 — 단일 진실 원천."""
    from collections import Counter
    values = _retry("통합 대장 읽기", lambda: qss.worksheet(MAIN_TAB).get_all_values())
    if not values:
        return {}, {}, {}
    ix = {h.strip(): i for i, h in enumerate(values[0])}
    d_i, r_i, p_i = ix.get("시행일자"), ix.get("연관도"), ix.get("우대여부")
    hi, si, pr = Counter(), Counter(), Counter()
    for row in values[1:]:
        d8 = norm_date(row[d_i]) if d_i is not None and d_i < len(row) else ""
        if len(d8) != 8:
            continue
        rel = str(row[r_i]).strip() if r_i is not None and r_i < len(row) else ""
        if rel == "연관높음":
            hi[d8] += 1
        elif rel == "단순관련":
            si[d8] += 1
        if p_i is not None and p_i < len(row) and str(row[p_i]).strip() == "O":
            pr[d8] += 1
    return hi, si, pr


def _pick_key(keys, must=(), ban=()):
    """헤더 표기 차이 대응: 조건에 맞는 첫 키를 고른다."""
    for k in keys:
        s = str(k)
        if all(m in s for m in must) and not any(b in s for b in ban):
            return k
    return None


def read_monitor_rows(ws):
    """monitor 총괄 → [(disp 'YYYY-MM-DD', date8, 총, 연관높음, 단순관련)] + 제외 통계"""
    records = _retry("monitor 총괄 읽기", lambda: ws.get_all_records())
    if not records:
        return [], {"빈 시트": 1}
    keys = list(records[0].keys())
    k_date = _pick_key(keys, must=("일자",)) or _pick_key(keys, must=("날짜",))
    k_total = _pick_key(keys, must=("검토",)) or _pick_key(keys, must=("총",), ban=("상태",))
    print(f"  🔎 monitor 헤더 매핑: 일자='{k_date}' 총='{k_total}' (연관/단순/우대는 대장에서 직접 카운트)")
    if not k_date or not k_total:
        print("⛔ monitor 총괄에서 일자/총검토 칸을 찾지 못했습니다."); sys.exit(1)

    rows, skip = [], {"날짜 인식 불가": 0, "건수 없음(실패일 등)": 0}
    for r in records:
        d8 = norm_date(r.get(k_date, ""))
        if len(d8) != 8:
            skip["날짜 인식 불가"] += 1
            continue
        try:
            total = int(float(r.get(k_total) or ""))
        except (ValueError, TypeError):
            skip["건수 없음(실패일 등)"] += 1
            continue

        disp = f"{d8[:4]}-{d8[4:6]}-{d8[6:]}"
        rows.append((disp, d8, total))
    rows.sort(key=lambda x: x[1])
    return rows, skip


def run(apply: bool):
    print("=" * 62)
    print(f"🗂️ 총괄현황표 백필 v2 (총검토=monitor / 관련건수=대장 직접카운트) — {'★실제 반영★' if apply else '미리보기만'}")
    print("=" * 62)
    since = input("시작 월 YYYYMM (엔터 = 전체 이력): ").strip()
    if since and not re.fullmatch(r"\d{6}", since):
        print("⛔ 형식이 다릅니다. 예: 202601"); return

    mss = _open_ss("MONITOR_SHEET")
    qss = _open_ss("QRADAR_SHEET")
    src_rows, skip = read_monitor_rows(mss.worksheet(TAB))
    if since:
        src_rows = [r for r in src_rows if r[1] >= since + "01"]

    print("  📊 통합 대장에서 날짜별 연관/단순/우대 직접 카운트 중…")
    hi_c, si_c, pr_c = count_from_ledger(qss)

    qws = qss.worksheet(TAB)
    qvals = _retry("Q-RADAR 총괄 읽기", lambda: qws.get_all_values())
    header = qvals[0] if qvals else []
    idx = {h.strip(): i for i, h in enumerate(header)}
    st_i = idx.get("모니터링 상태", 5)
    lg_i = idx.get("실행 로그 및 비고", 6)
    c_tot = idx.get("총 검토건수", 1)
    c_hi = idx.get("연관높음", 2)
    c_si = idx.get("단순관련", 3)
    c_pr = idx.get("우대건수", 4)
    existing = {}
    for rn, row in enumerate(qvals[1:], start=2):
        if row and str(row[0]).strip():
            existing[str(row[0]).strip()] = (rn, row)

    inserts, fills, keeps = [], [], []
    for disp, d8, total in src_rows:
        counts = [total, hi_c.get(d8, 0), si_c.get(d8, 0), pr_c.get(d8, 0)]
        if disp not in existing:
            inserts.append((disp, counts))
        else:
            rn, row = existing[disp]
            cur = [str(row[c]).strip() if c < len(row) else "" for c in (c_tot, c_hi, c_si, c_pr)]
            if all(v == "" for v in cur):
                fills.append((rn, disp, counts,
                              str(row[st_i]).strip() if st_i < len(row) else ""))
            else:
                keeps.append((disp, "건수 이미 존재 → 보존"))

    from collections import Counter
    monthly = Counter(d[:7] for d, _ in inserts) + Counter(d[:7] for _, d, _, _ in fills)
    print(f"  📥 monitor 검토일 {len(src_rows)}일 → 신규 삽입 {len(inserts)} / 기존행 소생 {len(fills)} / 보존 {len(keeps)}")
    for k, v in skip.items():
        if v:
            print(f"     · 제외 — {k}: {v}")
    if fills:
        print("  ♻️ 소생 대상:", ", ".join(f"{d}({st or '상태없음'})" for _rn, d, _c, st in fills[:4]))
    if monthly:
        print("  📅 월별:", ", ".join(f"{m} {n}일" for m, n in sorted(monthly.items())))

    from openpyxl import Workbook
    wb = Workbook()
    w1 = wb.active; w1.title = "신규삽입"
    w1.append(["시행일자", "총 검토건수", "연관높음(대장)", "단순관련(대장)", "우대건수(대장)", "상태(기록될 값)"])
    for d, c in inserts:
        w1.append([d] + c + [STAMP])
    w2 = wb.create_sheet("기존행소생")
    w2.append(["행번호", "시행일자", "채울 총검토", "연관", "단순", "우대", "기존 상태(불가침)"])
    for rn, d, c, st in fills:
        w2.append([rn, d] + c + [st])
    w3 = wb.create_sheet("보존")
    w3.append(["시행일자", "사유"])
    for d, why in keeps:
        w3.append([d, why])
    w4 = wb.create_sheet("월별요약"); w4.append(["월", "채움 일수"])
    for m in sorted(monthly):
        w4.append([m, monthly[m]])
    w5 = wb.create_sheet("통계")
    for k, v in [("monitor 검토일", len(src_rows)), ("신규 삽입", len(inserts)),
                 ("기존행 소생", len(fills)), ("보존", len(keeps))] + list(skip.items()):
        w5.append([k, v])
    try:
        wb.save(PREVIEW_PATH); print(f"  💾 미리보기: {PREVIEW_PATH}")
    except PermissionError:
        alt = str(HERE / f"총괄백필_미리보기_{time.strftime('%H%M%S')}.xlsx")
        wb.save(alt); print(f"  💾 미리보기: {alt}")

    if not apply:
        print("\n✅ 미리보기만 생성. 시트 확인 후 2번(반영)을 실행하세요."); return
    if not inserts and not fills:
        print("  반영할 것이 없습니다."); return

    def col(i):
        s = ""
        n = i + 1
        while n:
            n, r = divmod(n - 1, 26); s = chr(65 + r) + s
        return s

    if fills:
        updates = []
        for rn, _d, c, _st in fills:
            for ci, v in zip((c_tot, c_hi, c_si, c_pr), c):
                updates.append({"range": f"{col(ci)}{rn}", "values": [[v]]})
        for i in range(0, len(updates), 400):
            chunk = updates[i:i + 400]
            _retry(f"기존행 소생 기록({i + 1}~)", lambda c=chunk: qws.batch_update(c, value_input_option="RAW"))
        print(f"    ♻️ 기존행 소생 {len(fills)}행 (건수만 채움 — 상태·로그 불가침)")

    if inserts:
        width = max(len(header), st_i + 1, lg_i + 1, c_pr + 1)
        rows = []
        for d, c in inserts:
            row = [""] * width
            row[0] = d
            for ci, v in zip((c_tot, c_hi, c_si, c_pr), c):
                row[ci] = v
            row[st_i] = STAMP
            row[lg_i] = LOG_MSG
            rows.append(row)
        for i in range(0, len(rows), 200):
            chunk = rows[i:i + 200]
            _retry(f"신규 기록({i + 1}~)", lambda c=chunk: qws.append_rows(c, value_input_option="RAW"))
            print(f"    ➕ 신규 기록 {min(i + 200, len(rows))}/{len(rows)}행")
            time.sleep(0.5)

    print(f"\n🎉 백필 v2 완료: 신규 {len(inserts)} + 소생 {len(fills)} (보존 {len(keeps)})")
    print("   · 관련 건수는 전부 '지금의 대장'에서 파생 — 흡수·재분석 이력이 반영된 현재 진실")
    print("   👉 브리핑을 재실행하면 해당 월 KPI가 실수치로 나옵니다.")


if __name__ == "__main__":
    args = [a.lower() for a in sys.argv[1:]]
    if args:
        _preflight() and run(apply=("apply" in args)); sys.exit(0)
    print("=" * 62)
    print("🗂️ Q-RADAR 총괄현황표 백필 — 쉬운 모드")
    print("   (monitor의 일자별 검토 이력을 새 시트로 이관)")
    print("=" * 62)
    if not _preflight():
        sys.exit(0)
    print("\n무엇을 할까요?")
    print("  1) 미리보기 — 몇 행이 이관될지 목록만 (안 바꿈)")
    print("  2) 실제 반영 — Q-RADAR 총괄현황표에 🔵 도장으로 추가")
    print("  0) 종료")
    choice = input("\n번호를 입력하고 Enter: ").strip()
    if choice == "1":
        run(apply=False)
    elif choice == "2":
        confirm = input("⚠️ 새 시트 총괄현황표에 행이 추가됩니다. 진행하려면 '반영' 입력: ").strip()
        if confirm == "반영":
            run(apply=True)
        else:
            print("취소했습니다.")
    else:
        print("종료합니다.")
