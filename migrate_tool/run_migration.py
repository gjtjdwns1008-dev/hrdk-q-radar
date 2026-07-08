# -*- coding: utf-8 -*-
"""
run_migration.py — monitor + RADAR → Q-RADAR 통합 대장 이관 도구
=====================================================================
★미리보기 우선 원칙★
  1) 기본 실행 = 미리보기: 아무것도 안 바꾸고 '이관_미리보기.xlsx'만 생성
  2) 사람이 미리보기를 눈으로 확인
  3) APPLY=1 로 실행해야만 새 Q-RADAR 시트에 실제 반영

[이관 규칙 — 설계서 v1.1 §5]
  · 중복판정 키 = 법령명(공백정돈) | 시행일자(8자리)
  · 중복 시: RADAR 행 유지(우대 분석이 정밀) + monitor의 4칸만 보강
      (개정유형 / 주요 제·개정내용 / 활용도_구분 / 활용도_상세)
  · RADAR 기존 MST_ID 그대로, monitor 단독 행은 max+1부터 연번
  · monitor 단독 행의 우대여부 = 빈칸 (우대 분석을 안 했으므로 '모름'을 정직하게)
  · RADAR 행의 우대여부 = Track/우대분류 코드로 판정 (Z+Ⅳ-0+기타 → 빈칸, 그 외 → O)
  · Track 코드 라벨 정돈: 'B (영업요건형)' → 'B'  (새 brain 출력과 표기 통일)
  · 알 수 없는 칸의 수기 값 발견 시 → 검토사유에 [이관메모]로 보존 + 경고 목록

[사용법]
  ▶ 미리보기(라이브 시트): python run_migration.py live          ← Windows/맥 공통
  ▶ 실제 반영:            python run_migration.py live apply
  ▶ 미리보기(로컬 xlsx):   python run_migration.py               (.env에 xlsx 경로 지정 시)
  ※ bash 사용자는 SOURCE=live APPLY=1 환경변수 방식도 그대로 동작
     (부속 탭 복사 — 자격명칭최신화·우대사항_대장·보류목록 — 은 APPLY 때 함께)
"""
import os
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent

def _load_env_file():
    """migrate_tool/.env 를 읽어 dict로 (없으면 빈 dict) — Windows에서도 동작"""
    env = {}
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env

_ENV = _load_env_file()

# ══════════════ 여기만 고치세요 ══════════════
# [로컬 미리보기용] 두 시스템의 xlsx 내보내기 파일 경로 (.env에 적어도 됨)
MONITOR_XLSX = os.environ.get("MONITOR_XLSX", _ENV.get("MONITOR_XLSX",
               "/mnt/user-data/uploads/국가기술자격_법령_모니터링_Master_DB.xlsx"))
RADAR_XLSX   = os.environ.get("RADAR_XLSX", _ENV.get("RADAR_XLSX",
               "/mnt/user-data/uploads/HRDK_LAW-RADAR.xlsx"))

# [라이브용] 이 폴더에 gcp-key.json 과 .env(RADAR_SHEET / MONITOR_SHEET / QRADAR_SHEET) 를 두세요
SOURCE = os.environ.get("SOURCE", "local").lower()      # local | live
APPLY  = os.environ.get("APPLY", "").strip() == "1"      # 1 이어야만 실제 반영

# ★명령줄 인자 지원 — Windows PowerShell엔 `SOURCE=live python …` 문법이 없으므로:
#     python run_migration.py live          ← 라이브 시트 기준 미리보기
#     python run_migration.py live apply    ← 실제 반영 (빈 대장에만)
_args = [a.strip().lower() for a in sys.argv[1:]]
if "live" in _args:
    SOURCE = "live"
if "local" in _args:
    SOURCE = "local"
if "apply" in _args:
    APPLY = True

PREVIEW_PATH = str(HERE / "이관_미리보기.xlsx")
# ═════════════════════════════════════════════

# 통합 대장 24칸 (config.COLUMNS와 동일 — 도구 단독 실행 가능하게 여기 복제)
COLUMNS = [
    "MST_ID", "시행일자", "소관부처", "법령명", "개정유형", "연관도", "우대여부",
    "관련 종목", "주요 제·개정내용", "활용도_구분", "활용도_상세", "조문 요약",
    "우대분류", "Track1_취급유형", "Track1_위험도", "Track2_효용코드", "중처법대상",
    "상세 분석 결과", "근거조문", "AI신뢰도", "검토필요", "검토사유",
    "조문별 다이렉트 링크", "워크넷 실시간 구인건수",
]

MONITOR_MAP = {  # monitor 헤더 → 통합 칸
    "시행일자": "시행일자", "소관부처": "소관부처", "법령명": "법령명",
    "개정유형": "개정유형", "주요 제·개정내용": "주요 제·개정내용",
    "법령 관련 국가기술자격 종목": "관련 종목",
    "활용도 분석 구분": "활용도_구분", "활용도 분석 상세": "활용도_상세",
    "근거조문": "근거조문", "AI신뢰도": "AI신뢰도",
    "검토필요": "검토필요", "검토사유": "검토사유",
    "조문별 다이렉트 링크": "조문별 다이렉트 링크",
}
RADAR_MAP = {   # RADAR 헤더 → 통합 칸 (연관성_판별→연관도)
    "MST_ID": "MST_ID", "시행일자": "시행일자", "소관부처": "소관부처", "법령명": "법령명",
    "연관성_판별": "연관도", "관련 종목": "관련 종목", "조문 요약": "조문 요약",
    "우대분류": "우대분류", "Track1_취급유형": "Track1_취급유형",
    "Track1_위험도": "Track1_위험도", "Track2_효용코드": "Track2_효용코드",
    "중처법대상": "중처법대상", "상세 분석 결과": "상세 분석 결과",
    "근거조문": "근거조문", "AI신뢰도": "AI신뢰도", "검토필요": "검토필요",
    "검토사유": "검토사유", "조문별 다이렉트 링크": "조문별 다이렉트 링크",
    "워크넷 실시간 구인건수": "워크넷 실시간 구인건수",
}
ENRICH_COLS = ["개정유형", "주요 제·개정내용", "활용도_구분", "활용도_상세"]  # 중복 시 monitor→RADAR 보강 4칸
REAL_PREF = {"의무고용", "직무권한부여", "인사우대", "시험면제"}


# ── 유틸 ──────────────────────────────────────────────
def norm_law(s):
    return re.sub(r"\s+", " ", str(s or "").strip())

def norm_date(v):
    d = "".join(ch for ch in str(v or "") if ch.isdigit())
    return d[:8] if len(d) >= 8 else ""

def strip_code(v):
    """'B (영업요건형)' → 'B', 'Ⅳ-0 (제외)' → 'Ⅳ-0' — 새 brain 표기와 통일"""
    return str(v or "").strip().split(" ")[0].split("(")[0].strip()

def key_of(law, date):
    return f"{norm_law(law)}|{norm_date(date)}"


def read_tab_local(xlsx, tab):
    import openpyxl, warnings; warnings.filterwarnings("ignore")
    if not os.path.exists(xlsx):
        print(f"⛔ 로컬 xlsx를 찾을 수 없습니다: {xlsx}")
        print("   이 경로는 개발용 기본값입니다. 당신 PC에서는 둘 중 하나로 실행하세요:")
        print("   ① 라이브 시트에서 직접 읽기(권장):  python run_migration.py live")
        print("   ② 로컬 xlsx로 하기: .env에 RADAR_XLSX / MONITOR_XLSX 경로를 지정")
        sys.exit(1)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        return [], []
    it = wb[tab].iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    rows = [["" if c is None else str(c) for c in r] for r in it]
    return hdr, rows


def read_tab_live(sheet_env, tab):
    import json, gspread
    from oauth2client.service_account import ServiceAccountCredentials
    if not (HERE / "gcp-key.json").exists():
        print("⛔ migrate_tool 폴더에 gcp-key.json(서비스계정 키)이 없습니다.")
        print("   RADAR에서 쓰던 키 파일을 복사해 오고, 새 Q-RADAR 시트에도 그 서비스계정")
        print("   이메일을 편집자로 공유했는지 확인하세요.")
        sys.exit(1)
    key = _ENV.get(sheet_env, "")
    if not key:
        print(f"⛔ .env에 {sheet_env} 값이 없습니다.")
        print("   env.example을 복사해 .env를 만들고 세 시트 URL을 채워주세요.")
        sys.exit(1)
    m = re.search(r"/d/([A-Za-z0-9_-]+)", key)
    if m:
        key = m.group(1)
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        json.loads((HERE / "gcp-key.json").read_text(encoding="utf-8"), strict=False),
        ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"])
    ss = gspread.authorize(creds).open_by_key(key)
    try:
        vals = ss.worksheet(tab).get_all_values()
    except Exception:
        return [], []
    if not vals:
        return [], []
    return [h.strip() for h in vals[0]], vals[1:]


def rows_to_dicts(hdr, rows, mapping):
    """헤더 매핑으로 dict 변환 + 미지 칸의 수기 값 수집. 반환: (dict목록, 빈법령명_스킵수)
    ※ 구글시트 xlsx 내보내기는 빈 그리드 행까지 포함하므로 스킵 수가 큰 것이 정상."""
    known_idx = {h: i for i, h in enumerate(hdr) if h in mapping}
    unknown_idx = [(i, h) for i, h in enumerate(hdr) if h and h not in mapping]
    out = []
    skipped_empty = 0
    for r in rows:
        law = r[known_idx.get("법령명", 999)] if known_idx.get("법령명", 999) < len(r) else ""
        if not str(law).strip():
            skipped_empty += 1
            continue
        d = {}
        for h, i in known_idx.items():
            d[mapping[h]] = str(r[i]).strip() if i < len(r) else ""
        memos = []
        for i, h in unknown_idx:
            if i < len(r) and str(r[i]).strip():
                memos.append(f"{h}:{str(r[i]).strip()}")
        if memos:
            memo = "[이관메모] " + " / ".join(memos)
            d["검토사유"] = (d.get("검토사유", "") + " " + memo).strip()
            d["_memo"] = memo
        out.append(d)
    return out, skipped_empty


def decide_preferred(row):
    """RADAR 행의 우대여부 판정: Z+Ⅳ-0+기타 → '', 그 외 → 'O'"""
    t1 = strip_code(row.get("Track1_취급유형", ""))
    t2 = strip_code(row.get("Track2_효용코드", ""))
    pf = str(row.get("우대분류", "")).strip()
    if pf in REAL_PREF:
        return "O"
    if t1 and t1 != "Z":
        return "O"
    if t2 and t2 != "Ⅳ-0":
        return "O"
    return ""


def _preflight():
    """실행 전 준비물 점검 — 부족하면 뭘 해야 하는지 한국어로 안내"""
    ok = True
    if not (HERE / ".env").exists():
        if (HERE / "env.example").exists():
            (HERE / ".env").write_text((HERE / "env.example").read_text(encoding="utf-8"), encoding="utf-8")
            print("📄 .env 파일이 없어서 방금 만들어 뒀습니다 (env.example 복사본).")
        else:
            print("📄 .env 파일이 없습니다.")
        print("   👉 메모장으로  migrate_tool 폴더의 .env  를 열어 세 시트 URL을 붙여넣어 주세요.")
        print("      (구글시트를 브라우저로 열고 주소창의 URL 전체를 복사하면 됩니다)")
        ok = False
    else:
        env = _load_env_file()
        for k in ("RADAR_SHEET", "MONITOR_SHEET", "QRADAR_SHEET"):
            v = env.get(k, "")
            if not v or "여기에" in v:
                print(f"   ⚠️ .env의 {k} 칸이 아직 비어있어요 → 메모장으로 열어 해당 시트 URL을 붙여넣어 주세요.")
                ok = False
    if not (HERE / "gcp-key.json").exists():
        print("🔑 gcp-key.json(서비스계정 열쇠)이 이 폴더에 없습니다.")
        print("   👉 HRDK-LAW-RADAR의 local_backfill 폴더에서 쓰던 gcp-key.json을 이 폴더로 복사하세요.")
        print("   👉 그리고 새 Q-RADAR 시트의 [공유]에 그 서비스계정 이메일이 '편집자'로 있는지 확인!")
        ok = False
    if not ok:
        print("\n준비물을 채운 뒤, 다시 실행해 주세요:  python run_migration.py")
    return ok


def main():
    print("=" * 62)
    print(f"🚚 Q-RADAR 이관 도구 — 소스:{SOURCE} / {'★실제 반영★' if APPLY else '미리보기만'}")
    print("=" * 62)

    # ── 1) 읽기 ──
    if SOURCE == "live":
        r_hdr, r_rows = read_tab_live("RADAR_SHEET", "국가기술자격 관련법령")
        mh_hdr, mh_rows = read_tab_live("MONITOR_SHEET", "연관 높은 법령")
        ms_hdr, ms_rows = read_tab_live("MONITOR_SHEET", "국가기술자격 관계 법령(단순 관련)")
    else:
        r_hdr, r_rows = read_tab_local(RADAR_XLSX, "국가기술자격 관련법령")
        mh_hdr, mh_rows = read_tab_local(MONITOR_XLSX, "연관 높은 법령")
        ms_hdr, ms_rows = read_tab_local(MONITOR_XLSX, "국가기술자격 관계 법령(단순 관련)")
    print(f"  📥 RADAR {len(r_rows)}행 / monitor 연관높음 {len(mh_rows)}행 / 단순관련 {len(ms_rows)}행")

    radar, r_skip = rows_to_dicts(r_hdr, r_rows, RADAR_MAP)
    mon_h, mh_skip = rows_to_dicts(mh_hdr, mh_rows, MONITOR_MAP)
    mon_s, ms_skip = rows_to_dicts(ms_hdr, ms_rows, MONITOR_MAP)
    print(f"  🧹 실데이터: RADAR {len(radar)}행 / monitor 연관높음 {len(mon_h)}행 / 단순 {len(mon_s)}행 "
          f"(빈 그리드행 스킵: {r_skip}/{mh_skip}/{ms_skip} — 구글시트 내보내기 특성, 정상)")

    warnings_list = []

    # ── 2) RADAR: 동일키(법령명|시행일자) 그룹 → ★대표 1행으로 통합 ──
    #    대표행 = ①우대O 우선 → ②근거조문이 가장 긴(포괄적) 행 → ③MST_ID 낮은 행
    #    (같은 법령·같은 시행일을 여러 시기에 재분석한 '이력 지층'을 최선의 한 행으로)
    def _id_num(r):
        m = re.search(r"(\d+)$", str(r.get("MST_ID", "")))
        return int(m.group(1)) if m else 10 ** 9

    def _art_count(r):
        """근거조문의 조문 개수 (콤마 항목 수) — '포괄성'의 진짜 지표.
        텍스트 길이는 조문명을 길게 쓴 세대가 이기는 착시가 있어 2차 기준으로만."""
        s = str(r.get("근거조문", "")).strip()
        return len([t for t in s.split(",") if t.strip()]) if s else 0

    max_id = 0
    for d in radar:
        for c in ("Track1_취급유형", "Track1_위험도", "Track2_효용코드"):
            d[c] = strip_code(d.get(c, ""))
        d["연관도"] = str(d.get("연관도", "")).strip()
        d["우대여부"] = decide_preferred(d)
        if not norm_date(d.get("시행일자")):
            warnings_list.append(("RADAR", d.get("MST_ID", ""), d.get("법령명", ""), "시행일자 형식 이상"))
        mid = str(d.get("MST_ID", "")).strip()
        if mid.startswith("HRDK-L-"):
            try:
                max_id = max(max_id, int(mid.split("-")[-1]))
            except Exception:
                pass

    r_groups = {}
    for d in radar:
        r_groups.setdefault(key_of(d.get("법령명"), d.get("시행일자")), []).append(d)

    unified = {}
    consolidation_log = []
    absorbed_rows = 0
    for k, grp in r_groups.items():
        if len(grp) == 1:
            unified[k] = grp[0]
            continue
        rep = min(grp, key=lambda r: (0 if r.get("우대여부") == "O" else 1,
                                      -_art_count(r),
                                      -len(str(r.get("근거조문", ""))), _id_num(r)))
        absorbed = [str(x.get("MST_ID", "")) for x in grp if x is not rep]
        absorbed_rows += len(absorbed)
        same = len({(x.get("근거조문", ""), x.get("우대분류", ""),
                     x.get("Track1_취급유형", ""), x.get("Track2_효용코드", "")) for x in grp}) == 1
        tag = "동일내용 중복" if same else "조문범위 상이(이력 지층)"
        rep["검토사유"] = (str(rep.get("검토사유", "")) +
                        f" [이관통합:{tag}] {len(grp)}행→1, 흡수ID {', '.join(absorbed)}").strip()
        consolidation_log.append([rep.get("MST_ID", ""), rep.get("법령명", ""),
                                  norm_date(rep.get("시행일자")), len(grp), tag, ", ".join(absorbed)])
        unified[k] = rep

    # ── 3) monitor: 탭내·탭간 그룹 → 대표 1행(근거조문 긴 행), 그 후 RADAR에 병합/추가 ──
    stats = {"radar_raw": len(radar), "radar_uniq": len(unified),
             "absorbed": absorbed_rows, "consol_groups": len(consolidation_log),
             "mon_high": len(mon_h), "mon_simple": len(mon_s),
             "enriched": 0, "mon_new": 0, "mon_dup_skip": 0}
    merged_log = []

    mon_groups = {}
    for src_rows, rel in ((mon_h, "연관높음"), (mon_s, "단순관련")):
        for d in src_rows:
            d["_rel"] = rel
            mon_groups.setdefault(key_of(d.get("법령명"), d.get("시행일자")), []).append(d)

    for k, grp in mon_groups.items():
        stats["mon_dup_skip"] += len(grp) - 1
        rel = "연관높음" if any(x["_rel"] == "연관높음" for x in grp) else "단순관련"
        rep = max(grp, key=lambda r: (_art_count(r), len(str(r.get("근거조문", "")))))
        rep["연관도"] = rel
        rep.pop("_rel", None)
        if k in unified:                      # RADAR와 중복 → 4칸 보강
            base = unified[k]
            added = []
            for c in ENRICH_COLS:
                if str(rep.get(c, "")).strip() and not str(base.get(c, "")).strip():
                    base[c] = rep[c]
                    added.append(c)
            if rep.get("_memo"):
                base["검토사유"] = (base.get("검토사유", "") + " " + rep["_memo"]).strip()
            if added:
                stats["enriched"] += 1
                merged_log.append([base.get("MST_ID", ""), base.get("법령명", ""),
                                   norm_date(base.get("시행일자")), rel, ", ".join(added)])
        else:                                  # monitor 단독 → 신규 행
            rep["우대여부"] = ""              # 우대 분석 안 했으므로 '모름'
            unified[k] = rep
            stats["mon_new"] += 1

    # ── 4) MST_ID 신규 부여 + 정렬 ──
    rows = list(unified.values())
    rows.sort(key=lambda r: (norm_date(r.get("시행일자")) or "99999999",
                             str(r.get("MST_ID", "")) or "zzz"))
    for r in rows:
        if not str(r.get("MST_ID", "")).strip():
            max_id += 1
            r["MST_ID"] = f"HRDK-L-{max_id:04d}"

    total = len(rows)
    pref_cnt = sum(1 for r in rows if r.get("우대여부") == "O")
    stats.update({"total": total, "preferred": pref_cnt, "id_max": max_id})

    print(f"  🧬 RADAR 이력지층 통합: {stats['consol_groups']}그룹, {stats['absorbed']}행 흡수 "
          f"({stats['radar_raw']}→{stats['radar_uniq']}행)")
    print(f"  🔗 병합(보강) {stats['enriched']}건 / monitor 단독 신규 {stats['mon_new']}건 / "
          f"monitor 중복 정리 {stats['mon_dup_skip']}건")
    print(f"  📊 최종 통합 대장 {total}행 (우대 O = {pref_cnt}행, MST_ID ~{max_id:04d})")
    if warnings_list:
        print(f"  ⚠️ 경고 {len(warnings_list)}건 (미리보기 '경고' 시트 확인)")

    # ── 5) 미리보기 xlsx ──
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "통합대장(미리보기)"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    ws2 = wb.create_sheet("병합내역")
    ws2.append(["MST_ID", "법령명", "시행일자", "monitor측 연관도", "보강된 칸"])
    for m in merged_log:
        ws2.append(m)
    ws25 = wb.create_sheet("이관통합내역(RADAR)")
    ws25.append(["대표 MST_ID", "법령명", "시행일자", "그룹 행수", "성격", "흡수된 MST_ID"])
    for c in consolidation_log:
        ws25.append(c)
    ws3 = wb.create_sheet("통계")
    for k, v in [("RADAR 실데이터 행", stats["radar_raw"]),
                 ("RADAR 이력지층 통합", f"{stats['consol_groups']}그룹 / {stats['absorbed']}행 흡수"),
                 ("RADAR 고유(대표) 행", stats["radar_uniq"]),
                 ("monitor 연관높음", stats["mon_high"]),
                 ("monitor 단순관련", stats["mon_simple"]), ("병합(보강)", stats["enriched"]),
                 ("monitor 단독 신규", stats["mon_new"]), ("monitor 중복 정리", stats["mon_dup_skip"]),
                 ("최종 통합 행", stats["total"]), ("우대여부 O", stats["preferred"]),
                 ("MST_ID 최대", f"HRDK-L-{stats['id_max']:04d}"),
                 ("대표행 규칙", "①우대O 우선 ②조문 개수 많은 행 ③텍스트 긴 행 ④MST_ID 낮은 행 — 흡수 이력은 검토사유+통합내역 시트"),
                 ("공통 규칙", "중복키=법령명|시행일자, RADAR 우선+monitor 4칸 보강, Track 코드 라벨 제거")]:
        ws3.append([k, v])
    ws4 = wb.create_sheet("경고")
    ws4.append(["출처", "MST_ID", "법령명", "내용"])
    for w in warnings_list:
        ws4.append(list(w))
    wb.save(PREVIEW_PATH)
    print(f"  💾 미리보기 저장: {PREVIEW_PATH}")

    # ── 6) 실제 반영 (APPLY=1 + live에서만) ──
    if not APPLY:
        print("\n✅ 미리보기만 생성했습니다. (시트는 아무것도 바뀌지 않았어요)")
        print("   👉 위 경로의 '이관_미리보기.xlsx'를 열어 확인한 뒤,")
        print("      다시  python run_migration.py  실행 → 2번(실제 반영)을 고르세요.")
        return
    if SOURCE != "live":
        print("\n⛔ APPLY는 SOURCE=live 에서만 허용됩니다 (로컬 xlsx로는 반영 불가).")
        sys.exit(1)

    import json, gspread
    from oauth2client.service_account import ServiceAccountCredentials
    qkey = _ENV.get("QRADAR_SHEET", "")
    if not qkey:
        print("⛔ .env에 QRADAR_SHEET(새 시트 URL)가 없습니다 — 반영 중단.")
        sys.exit(1)
    m = re.search(r"/d/([A-Za-z0-9_-]+)", qkey)
    if m:
        qkey = m.group(1)
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        json.loads((HERE / "gcp-key.json").read_text(encoding="utf-8"), strict=False),
        ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"])
    gc = gspread.authorize(creds)
    q = gc.open_by_key(qkey)

    # 6-1) 통합 대장
    try:
        ws_main = q.worksheet("국가기술자격 관련법령")
    except gspread.WorksheetNotFound:
        ws_main = q.add_worksheet(title="국가기술자격 관련법령", rows=total + 50, cols=len(COLUMNS) + 2)
    existing = ws_main.get_all_values()
    if len(existing) > 1:
        print(f"  ⛔ 대상 시트 대장에 이미 {len(existing)-1}행 존재 — 안전을 위해 중단 (빈 대장에만 이관)")
        sys.exit(1)
    ws_main.clear()
    payload = [COLUMNS] + [[r.get(c, "") for c in COLUMNS] for r in rows]
    for i in range(0, len(payload), 500):
        ws_main.append_rows(payload[i:i + 500], value_input_option="RAW")
        print(f"    …대장 {min(i+500, len(payload))}/{len(payload)}행 적재")

    # 6-1.5) ★총괄현황표 이음새(기준선) — Q-RADAR 첫 실행이 '이관 최신일 +1'부터 분석하도록
    #   (이게 없으면 첫 실행이 '어제 하루만' 폴백 — 이관~가동 사이 며칠 뜨면 구멍 생김)
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    _max_date = max((norm_date(r.get("시행일자")) for r in rows if norm_date(r.get("시행일자"))), default="")
    if _max_date:
        try:
            ws_sum = q.worksheet("총괄현황표")
        except gspread.WorksheetNotFound:
            ws_sum = q.add_worksheet(title="총괄현황표", rows=400, cols=9)
            ws_sum.append_row(["시행일자", "총 검토건수", "연관높음", "단순관련", "우대건수",
                               "모니터링 상태", "실행 로그 및 비고"])
        if len(ws_sum.get_all_values()) <= 1:
            _stamp = _dt.now(_tz(_td(hours=9))).strftime("%m/%d %H:%M")
            _disp = f"{_max_date[:4]}-{_max_date[4:6]}-{_max_date[6:]}"
            ws_sum.append_row([_disp, "", "", "", "", f"{_stamp}🟢",
                               "이관 기준선 — 이 시행일자까지는 기존 두 시스템 분석분(이관)"])
            print(f"    🧷 이음새 설정: 마지막 성공일 = {_disp} → 다음 실행은 +1일부터 분석")

    # 6-2) 부속 탭 복사 (RADAR 라이브 → Q-RADAR): 자격명칭최신화 / 우대사항_대장 / 보류목록
    for tab in ("자격명칭최신화", "우대사항_대장", "보류목록"):
        hdr, body = read_tab_live("RADAR_SHEET", tab)
        if not hdr:
            print(f"    ⚠️ 원본 '{tab}' 없음 — 건너뜀")
            continue
        try:
            wt = q.worksheet(tab)
            if len(wt.get_all_values()) > 1:
                print(f"    ⏭️ '{tab}' 이미 데이터 있음 — 덮어쓰지 않음")
                continue
        except gspread.WorksheetNotFound:
            wt = q.add_worksheet(title=tab, rows=len(body) + 20, cols=len(hdr) + 2)
        wt.clear()
        data = [hdr] + body
        for i in range(0, len(data), 500):
            wt.append_rows(data[i:i + 500], value_input_option="RAW")
        print(f"    📋 '{tab}' 복사 완료 ({len(body)}행)")

    print("\n🎉 이관 반영 완료! 다음: Q-RADAR 첫 수동 실행으로 총괄현황표가 자동 생성되는지 확인하세요.")


if __name__ == "__main__":
    if not _args:   # ★인자 없이 실행 = 초심자 모드: 번호로 고르는 한국어 메뉴
        print("=" * 62)
        print("🚚 Q-RADAR 이관 도구 — 쉬운 모드")
        print("=" * 62)
        if not _preflight():
            sys.exit(0)
        print("\n무엇을 할까요?")
        print("  1) 미리보기  — 라이브 시트를 읽어 '이관_미리보기.xlsx'만 생성 (아무것도 안 바꿈, 안전)")
        print("  2) 실제 반영 — 새 Q-RADAR 시트에 이관 (빈 대장에만 들어가는 안전장치 있음)")
        print("  0) 종료")
        choice = input("\n번호를 입력하고 Enter: ").strip()
        if choice == "1":
            SOURCE = "live"
            APPLY = False
        elif choice == "2":
            SOURCE = "live"
            confirm = input("⚠️ 새 시트에 실제로 기록합니다. 진행하려면 '반영' 이라고 입력: ").strip()
            if confirm != "반영":
                print("취소했습니다. (아무 변화 없음)")
                sys.exit(0)
            APPLY = True
        else:
            print("종료합니다. (아무 변화 없음)")
            sys.exit(0)
    main()
